from flask import request, jsonify, session, send_file
from openpyxl import Workbook, load_workbook
from io import BytesIO
import datetime
import re
import database
from routes.utils import is_logged_in, require_privilege

def get_header_idx(header_map, *keys):
    for k in keys:
        if k is not None:
            clean_k = str(k).strip().lower()
            idx = header_map.get(clean_k)
            if idx is not None:
                return idx
                
    for k in keys:
        if k is not None:
            clean_k = str(k).strip().lower()
            for header_name, idx in header_map.items():
                if clean_k in header_name or header_name in clean_k:
                    return idx
    return None

def parse_date_robust(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime('%Y-%m-%d')
    
    date_str = str(date_val).strip()
    formats = [
        '%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y',
        '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y',
        '%Y.%m.%d', '%d.%m.%Y', '%m.%d.%Y'
    ]
    for fmt in formats:
        try:
            dt = datetime.datetime.strptime(date_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def register_excel_routes(app):
    @app.route('/api/expenses/import-template', methods=['GET'])
    def get_import_template():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        db_cols = database.get_excel_columns('expense')
        active_cols = [c for c in db_cols if c['is_enabled_import'] == 1]
        
        headers = [c['column_label'] for c in active_cols]
        ws.append(headers)
        
        placeholders = {
            'date': "2026-06-25",
            'category': "Food & Dining",
            'description': "Lunch with friends",
            'gateway': "GPay",
            'bank': "SBI",
            'source': "Salary",
            'method': "Debit",
            'amount': 12.50,
            'interest': 0.00,
            'status': "Paid"
        }
        placeholders2 = {
            'date': "2026-06-26",
            'category': "Shopping",
            'description': "Bought a laptop",
            'gateway': "Credit Card",
            'bank': "Kotak",
            'source': "Credit Card",
            'method': "Credit",
            'amount': 1200.00,
            'interest': 45.00,
            'status': "Unpaid"
        }
        
        row1 = [placeholders.get(c['column_key'], '') for c in active_cols]
        row2 = [placeholders2.get(c['column_key'], '') for c in active_cols]
        ws.append(row1)
        ws.append(row2)
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="spendsmart_import_template.xlsx"
        )

    @app.route('/api/expenses/export', methods=['GET'])
    @require_privilege('can_view')
    def export_expenses():
        month = request.args.get('month')
        year = request.args.get('year')
        
        expenses = database.get_expenses(
            session['user_id'],
            month=month,
            year=year
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        db_cols = database.get_excel_columns('expense')
        active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
        
        headers = [c['column_label'] for c in active_cols]
        ws.append(headers)
        
        for exp in expenses:
            row_data = []
            for col in active_cols:
                k = col['column_key']
                if k == 'date':
                    row_data.append(exp.get('date', ''))
                elif k == 'category':
                    row_data.append(exp.get('category', ''))
                elif k == 'description':
                    row_data.append(exp.get('description', ''))
                elif k == 'gateway':
                    row_data.append(exp.get('payment_type', ''))
                elif k == 'bank':
                    row_data.append(exp.get('bank_mode', ''))
                elif k == 'source':
                    row_data.append(exp.get('payment_category', ''))
                elif k == 'method':
                    row_data.append(exp.get('payment_method', 'Debit'))
                elif k == 'amount':
                    row_data.append(float(exp.get('amount', 0.0)))
                elif k == 'interest':
                    row_data.append(float(exp.get('interest', 0.0)))
                elif k == 'status':
                    row_data.append(exp.get('status', 'Paid'))
                else:
                    row_data.append(exp.get(k, ''))
            ws.append(row_data)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        download_name = f"expenses_export"
        if year:
            download_name += f"_{year}"
        if month:
            download_name += f"_{month}"
        download_name += f"_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name
        )

    @app.route('/api/expenses/import', methods=['POST'])
    @require_privilege('can_add')
    def import_expenses():
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
            
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                return jsonify({'error': 'The uploaded Excel file is empty'}), 400
                
            header_map = {}
            for i, val in enumerate(first_row):
                if val is not None:
                    header_map[str(val).strip().lower()] = i
                    
            db_cols = database.get_excel_columns('expense')
            enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
            
            col_date = get_header_idx(header_map, 'date') if 'date' in enabled_keys else None
            col_category = get_header_idx(header_map, 'category') if 'category' in enabled_keys else None
            col_amount = get_header_idx(header_map, 'amount') if 'amount' in enabled_keys else None
            
            missing_reqs = []
            if 'date' in enabled_keys and col_date is None: missing_reqs.append('Date')
            if 'category' in enabled_keys and col_category is None: missing_reqs.append('Category')
            if 'amount' in enabled_keys and col_amount is None: missing_reqs.append('Amount')
            
            if missing_reqs:
                return jsonify({
                    'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'
                }), 400
                
            col_desc = get_header_idx(header_map, 'description') if 'description' in enabled_keys else None
            col_gateway = get_header_idx(header_map, 'gateway', 'payment gateway', 'payment type', 'payment_type') if 'gateway' in enabled_keys else None
            col_bank = get_header_idx(header_map, 'bank', 'bank name', 'bank_name', 'bank mode', 'bank_mode') if 'bank' in enabled_keys else None
            col_source = get_header_idx(header_map, 'source', 'payment source', 'payment_source', 'payment category', 'payment_category') if 'source' in enabled_keys else None
            col_method = get_header_idx(header_map, 'method', 'payment method', 'payment_method') if 'method' in enabled_keys else None
            col_interest = get_header_idx(header_map, 'interest') if 'interest' in enabled_keys else None
            col_status = get_header_idx(header_map, 'status') if 'status' in enabled_keys else None
            
            dynamic_cols = {}
            for col in db_cols:
                k = col['column_key']
                if col['is_enabled_import'] == 1 and k not in ('date', 'category', 'description', 'gateway', 'bank', 'source', 'method', 'amount', 'interest', 'status'):
                    col_idx = get_header_idx(header_map, k, col['column_label'])
                    if col_idx is not None:
                        dynamic_cols[k] = col_idx
                        
            imported_count = 0
            skipped_count = 0
            
            for r_idx, row in enumerate(rows_iter, start=2):
                if not any(val is not None for val in row):
                    continue
                    
                date_val = row[col_date] if col_date is not None else None
                category_val = row[col_category] if col_category is not None else None
                amount_val = row[col_amount] if col_amount is not None else None
                
                if 'date' in enabled_keys and col_date is not None:
                    if date_val is None:
                        skipped_count += 1
                        continue
                    date_str = parse_date_robust(date_val)
                    if date_str is None:
                        skipped_count += 1
                        continue
                else:
                    date_str = datetime.date.today().strftime('%Y-%m-%d')
                    
                if 'category' in enabled_keys and col_category is not None:
                    if category_val is None:
                        skipped_count += 1
                        continue
                    category_val = str(category_val).strip()
                    if not category_val:
                        skipped_count += 1
                        continue
                else:
                    category_val = 'Other'
                    
                if 'amount' in enabled_keys and col_amount is not None:
                    if amount_val is None:
                        skipped_count += 1
                        continue
                    try:
                        amount = float(amount_val)
                        if amount <= 0:
                            skipped_count += 1
                            continue
                    except (ValueError, TypeError):
                        skipped_count += 1
                        continue
                else:
                    amount = 0.01
                    
                description = str(row[col_desc]).strip() if (col_desc is not None and row[col_desc] is not None) else ''
                gateway = str(row[col_gateway]).strip() if (col_gateway is not None and row[col_gateway] is not None) else ''
                bank = str(row[col_bank]).strip() if (col_bank is not None and row[col_bank] is not None) else ''
                source = str(row[col_source]).strip() if (col_source is not None and row[col_source] is not None) else ''
                method = str(row[col_method]).strip() if (col_method is not None and row[col_method] is not None) else 'Debit'
                
                if method.lower() in ('credit', 'c'):
                    method = 'Credit'
                else:
                    method = 'Debit'
                    
                interest = 0.0
                if col_interest is not None and row[col_interest] is not None:
                    try:
                        interest = float(row[col_interest])
                        if interest < 0:
                            interest = 0.0
                    except (ValueError, TypeError):
                        interest = 0.0
                        
                if method == 'Debit':
                    interest = 0.0
                    
                status_val = 'Paid'
                if col_status is not None and row[col_status] is not None:
                    status_raw = str(row[col_status]).strip().lower()
                    if status_raw in ('unpaid', 'u'):
                        status_val = 'Unpaid'
                        
                extra_fields = {}
                for k, col_idx in dynamic_cols.items():
                    val = row[col_idx]
                    extra_fields[k] = str(val).strip() if val is not None else ''

                database.add_expense(
                    session['user_id'], amount, category_val, description, date_str,
                    bank_mode=bank, payment_type=gateway, payment_category=source,
                    interest=interest, payment_method=method, status=status_val,
                    **extra_fields
                )
                imported_count += 1
                
            return jsonify({
                'success': True,
                'message': f'Import completed. {imported_count} expenses imported successfully. {skipped_count} invalid rows skipped.',
                'imported': imported_count,
                'skipped': skipped_count
            })
        except Exception as e:
            return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

    @app.route('/api/emis/import-template', methods=['GET'])
    def get_emi_import_template():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        db_cols = database.get_excel_columns('emi')
        active_cols = [c for c in db_cols if c['is_enabled_import'] == 1]
        
        headers = [c['column_label'] for c in active_cols]
        ws.append(headers)
        
        placeholders = {
            'name': "Car Loan EMI",
            'principal_amount': 15000.00,
            'interest_rate': 7.5,
            'tenure_months': 36,
            'emi_amount': 466.50,
            'start_date': "2026-01-01",
            'end_date': "2028-12-01",
            'due_date': "5",
            'payment_type': "Auto Debit",
            'payment_gateway': "NACH",
            'payment_bank': "HDFC"
        }
        
        row1 = [placeholders.get(c['column_key'], '') for c in active_cols]
        ws.append(row1)
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="emi_import_template.xlsx"
        )

    @app.route('/api/emis/export', methods=['GET'])
    @require_privilege('can_view')
    def export_emis():
        emis = database.get_emis(session['user_id'])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EMIs"
        
        db_cols = database.get_excel_columns('emi')
        active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
        
        headers = [c['column_label'] for c in active_cols]
        ws.append(headers)
        
        for emi in emis:
            row_data = []
            for col in active_cols:
                k = col['column_key']
                if k == 'name':
                    row_data.append(emi.get('name', ''))
                elif k == 'principal_amount':
                    row_data.append(float(emi.get('principal_amount', 0.0)))
                elif k == 'interest_rate':
                    row_data.append(float(emi.get('interest_rate', 0.0)))
                elif k == 'tenure_months':
                    row_data.append(int(emi.get('tenure_months', 0)))
                elif k == 'emi_amount':
                    row_data.append(float(emi.get('emi_amount', 0.0)))
                elif k == 'start_date':
                    row_data.append(emi.get('start_date', ''))
                elif k == 'end_date':
                    row_data.append(emi.get('end_date', ''))
                elif k == 'due_date':
                    row_data.append(emi.get('due_date', ''))
                elif k == 'payment_type':
                    row_data.append(emi.get('payment_type', ''))
                elif k == 'payment_gateway':
                    row_data.append(emi.get('payment_gateway', ''))
                elif k == 'payment_bank':
                    row_data.append(emi.get('payment_bank', ''))
                else:
                    row_data.append(emi.get(k, ''))
            ws.append(row_data)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        download_name = f"emis_export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name
        )

    @app.route('/api/emis/import', methods=['POST'])
    @require_privilege('can_add')
    def import_emis():
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
            
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                return jsonify({'error': 'The uploaded Excel file is empty'}), 400
                
            header_map = {}
            for i, val in enumerate(first_row):
                if val is not None:
                    header_map[str(val).strip().lower()] = i
                    
            db_cols = database.get_excel_columns('emi')
            enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
            
            missing_reqs = []
            for k, req in enabled_keys.items():
                if req:
                    label = next(c['column_label'] for c in db_cols if c['column_key'] == k)
                    if label.lower() not in header_map:
                        missing_reqs.append(label)
            if missing_reqs:
                return jsonify({'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'}), 400
                
            imported_count = 0
            user_id = session['user_id']
            
            for r_idx, row in enumerate(rows_iter, start=2):
                if not any(val is not None for val in row):
                    continue
                    
                data_dict = {}
                for col in db_cols:
                    k = col['column_key']
                    if col['is_enabled_import'] != 1:
                        continue
                    label = col['column_label']
                    idx = get_header_idx(header_map, label, k)
                    if idx is not None:
                        data_dict[k] = row[idx]
                
                name = str(data_dict.get('name') or '').strip()
                principal = float(data_dict.get('principal_amount') or 0.0)
                interest = float(data_dict.get('interest_rate') or 0.0)
                tenure = int(data_dict.get('tenure_months') or 12)
                emi_amount = float(data_dict.get('emi_amount') or 0.0)
                
                start_date_val = data_dict.get('start_date')
                end_date_val = data_dict.get('end_date')
                
                start_date = parse_date_robust(start_date_val)
                end_date = parse_date_robust(end_date_val)
                    
                due_date = str(data_dict.get('due_date') or '5').strip()
                payment_type = str(data_dict.get('payment_type') or 'Auto').strip()
                gateway = str(data_dict.get('payment_gateway') or '').strip()
                bank = str(data_dict.get('payment_bank') or '').strip()
                
                if not name or not start_date or not end_date or not due_date or not payment_type:
                    continue
                    
                extra_fields = {}
                for k in data_dict:
                    if k not in ('name', 'principal_amount', 'interest_rate', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'):
                        extra_fields[k] = str(data_dict[k]).strip() if data_dict[k] is not None else ''
                        
                database.add_emi(
                    user_id, name, principal, emi_amount, start_date, end_date, tenure, interest, due_date, payment_type, gateway, bank,
                    **extra_fields
                )
                imported_count += 1
                
            return jsonify({'success': True, 'message': f'Successfully imported {imported_count} EMIs.'})
        except Exception as e:
            return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

    @app.route('/api/admin/emis/export', methods=['GET'])
    @require_privilege('can_admin')
    def admin_export_emis():
        emis = database.get_all_emis()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "All EMIs"
        
        db_cols = database.get_excel_columns('emi')
        active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
        
        headers = ["Username"] + [c['column_label'] for c in active_cols]
        ws.append(headers)
        
        for emi in emis:
            row_data = [emi.get('username', '')]
            for col in active_cols:
                k = col['column_key']
                if k == 'name':
                    row_data.append(emi.get('name', ''))
                elif k == 'principal_amount':
                    row_data.append(float(emi.get('principal_amount', 0.0)))
                elif k == 'interest_rate':
                    row_data.append(float(emi.get('interest_rate', 0.0)))
                elif k == 'tenure_months':
                    row_data.append(int(emi.get('tenure_months', 0)))
                elif k == 'emi_amount':
                    row_data.append(float(emi.get('emi_amount', 0.0)))
                elif k == 'start_date':
                    row_data.append(emi.get('start_date', ''))
                elif k == 'end_date':
                    row_data.append(emi.get('end_date', ''))
                elif k == 'due_date':
                    row_data.append(emi.get('due_date', ''))
                elif k == 'payment_type':
                    row_data.append(emi.get('payment_type', ''))
                elif k == 'payment_gateway':
                    row_data.append(emi.get('payment_gateway', ''))
                elif k == 'payment_bank':
                    row_data.append(emi.get('payment_bank', ''))
                else:
                    row_data.append(emi.get(k, ''))
            ws.append(row_data)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        download_name = f"admin_emis_export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name
        )

    @app.route('/api/admin/emis/import', methods=['POST'])
    @require_privilege('can_admin')
    def admin_import_emis():
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
            
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                return jsonify({'error': 'The uploaded Excel file is empty'}), 400
                
            header_map = {}
            for i, val in enumerate(first_row):
                if val is not None:
                    header_map[str(val).strip().lower()] = i
                    
            db_cols = database.get_excel_columns('emi')
            enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
            
            if 'username' not in header_map:
                return jsonify({'error': 'Required column "Username" missing in Excel.'}), 400
                
            missing_reqs = []
            for k, req in enabled_keys.items():
                if req:
                    label = next(c['column_label'] for c in db_cols if c['column_key'] == k)
                    if label.lower() not in header_map:
                        missing_reqs.append(label)
            if missing_reqs:
                return jsonify({'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'}), 400
                
            imported_count = 0
            skipped_count = 0
            
            for r_idx, row in enumerate(rows_iter, start=2):
                if not any(val is not None for val in row):
                    continue
                    
                username = str(row[header_map["username"]] or '').strip()
                if not username:
                    skipped_count += 1
                    continue
                    
                user = database.get_user_by_username(username)
                if not user:
                    skipped_count += 1
                    continue
                    
                user_id = user['id']
                
                data_dict = {}
                for col in db_cols:
                    k = col['column_key']
                    if col['is_enabled_import'] != 1:
                        continue
                    label = col['column_label']
                    idx = get_header_idx(header_map, label, k)
                    if idx is not None:
                        data_dict[k] = row[idx]
                
                name = str(data_dict.get('name') or '').strip()
                principal = float(data_dict.get('principal_amount') or 0.0)
                interest = float(data_dict.get('interest_rate') or 0.0)
                tenure = int(data_dict.get('tenure_months') or 12)
                emi_amount = float(data_dict.get('emi_amount') or 0.0)
                
                start_date_val = data_dict.get('start_date')
                end_date_val = data_dict.get('end_date')
                
                start_date = parse_date_robust(start_date_val)
                end_date = parse_date_robust(end_date_val)
                    
                due_date = str(data_dict.get('due_date') or '5').strip()
                payment_type = str(data_dict.get('payment_type') or 'Auto').strip()
                gateway = str(data_dict.get('payment_gateway') or '').strip()
                bank = str(data_dict.get('payment_bank') or '').strip()
                
                if not name or not start_date or not end_date or not due_date or not payment_type:
                    skipped_count += 1
                    continue
                    
                extra_fields = {}
                for k in data_dict:
                    if k not in ('name', 'principal_amount', 'interest_rate', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'):
                        extra_fields[k] = str(data_dict[k]).strip() if data_dict[k] is not None else ''
                        
                database.add_emi(
                    user_id, name, principal, emi_amount, start_date, end_date, tenure, interest, due_date, payment_type, gateway, bank,
                    **extra_fields
                )
                imported_count += 1
                
            msg = f'Successfully imported {imported_count} EMIs.'
            if skipped_count > 0:
                msg += f' Skipped {skipped_count} rows due to invalid/missing username or data.'
                
            return jsonify({'success': True, 'message': msg})
        except Exception as e:
            return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

    @app.route('/api/admin/excel-columns', methods=['GET'])
    def admin_get_excel_columns():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        target_type = request.args.get('target_type', 'expense')
        cols = database.get_excel_columns(target_type)
        return jsonify(cols)

    @app.route('/api/admin/excel-columns/toggle', methods=['POST'])
    def admin_toggle_excel_column():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        data = request.get_json() or {}
        target_type = data.get('target_type', 'expense')
        
        privilege = 'Expense Columns List'
        if target_type == 'emi':
            privilege = 'EMI Columns List'
        elif target_type == 'excel':
            privilege = 'Excel Import & Export Columns'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'edit'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
        column_key = data.get('column_key')
        is_enabled = data.get('is_enabled')
        type_key = data.get('type_key')
        target_type = data.get('target_type', 'expense')
        
        if column_key is None or is_enabled is None or type_key not in ('import', 'export'):
            return jsonify({'error': 'Column key, is_enabled, and valid type_key ("import" or "export") are required.'}), 400
            
        db_cols = database.get_excel_columns(target_type)
        target_col = next((c for c in db_cols if c['column_key'] == column_key), None)
        if not target_col:
            return jsonify({'error': 'Column not found.'}), 404
            
        database.update_excel_column_status(column_key, type_key, int(is_enabled), target_type)
        return jsonify({'success': True, 'message': 'Column status updated successfully.'})

    @app.route('/api/admin/excel-columns/save-all', methods=['POST'])
    def admin_save_all_excel_columns():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        data = request.get_json() or {}
        columns_data = data.get('columns', [])
        
        # Check target_type of the first item to determine privilege
        target_type = 'expense'
        if columns_data:
            target_type = columns_data[0].get('target_type', 'expense')
            
        privilege = 'Expense Columns List'
        if target_type == 'emi':
            privilege = 'EMI Columns List'
        elif target_type == 'excel':
            privilege = 'Excel Import & Export Columns'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'edit'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
        columns_data = data.get('columns', [])
        type_key = data.get('type_key', 'import')
        
        if type_key not in ('import', 'export'):
            return jsonify({'error': 'Invalid type_key ("import" or "export")'}), 400
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        field = 'is_enabled_import' if type_key == 'import' else 'is_enabled_export'
        
        for col in columns_data:
            col_key = col.get('column_key')
            target_type = col.get('target_type', 'expense')
            display_order = col.get('display_order', 0)
            is_enabled = col.get('is_enabled', 1)
            
            cursor.execute(
                f"UPDATE excel_columns SET display_order = ?, {field} = ? WHERE column_key = ? AND target_type = ?",
                (int(display_order), int(is_enabled), col_key, target_type)
            )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'All configurations saved successfully.'})

    @app.route('/api/admin/excel-columns/update-order', methods=['POST'])
    def admin_update_excel_column_order():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        data = request.get_json() or {}
        target_type = data.get('target_type', 'expense')
        
        privilege = 'Expense Columns List'
        if target_type == 'emi':
            privilege = 'EMI Columns List'
        elif target_type == 'excel':
            privilege = 'Excel Import & Export Columns'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'edit'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
        column_key = data.get('column_key')
        display_order = data.get('display_order')
        target_type = data.get('target_type', 'expense')
        
        if column_key is None or display_order is None:
            return jsonify({'error': 'Column key and display_order are required.'}), 400
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE excel_columns SET display_order = ? WHERE column_key = ? AND target_type = ?",
            (int(display_order), column_key, target_type)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Column order updated successfully.'})

    @app.route('/api/admin/excel-columns/create', methods=['POST'])
    def admin_create_excel_column():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        import sqlite3
        data = request.get_json() or {}
        target_type = data.get('target_type', 'expense').strip().lower()
        
        privilege = 'Add Custom Column for Expenses'
        if target_type == 'emi':
            privilege = 'Add Custom Column for EMIs'
        elif target_type == 'excel':
            privilege = 'Add Custom Column'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'add'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
        column_key = data.get('column_key', '').strip().lower()
        column_label = data.get('column_label', '').strip()
        target_type = data.get('target_type', 'expense').strip().lower()
        is_required = int(data.get('is_required', 0))
        is_enabled_import = int(data.get('is_enabled_import', 1))
        is_enabled_export = int(data.get('is_enabled_export', 1))
        display_order = int(data.get('display_order', 0))
        parent_column_key = data.get('parent_column_key', '').strip() or None
        parent_trigger_value = data.get('parent_trigger_value', '').strip() or None
        
        if not column_key or not column_label or target_type not in ('expense', 'emi'):
            return jsonify({'error': 'Column key, label, and valid target_type ("expense" or "emi") are required.'}), 400
            
        import re
        if not re.match(r'^[a-z0-9_]+$', column_key):
            return jsonify({'error': 'Column key must only contain lowercase alphanumeric characters and underscores.'}), 400
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        try:
            table_name = 'expenses' if target_type == 'expense' else 'emis'
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = [row['name'] for row in cursor.fetchall()]
            if column_key not in columns:
                cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_key} TEXT")
                
            cursor.execute(
                'INSERT INTO excel_columns (column_key, column_label, is_enabled_import, is_enabled_export, is_required, target_type, display_order, parent_column_key, parent_trigger_value) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (column_key, column_label, is_enabled_import, is_enabled_export, is_required, target_type, display_order, parent_column_key, parent_trigger_value)
            )
            conn.commit()
            return jsonify({'success': True, 'message': f'Column registered and added to {table_name} table successfully.'})
        except sqlite3.IntegrityError:
            return jsonify({'error': 'Column key already exists for this target type.'}), 409
        except Exception as e:
            return jsonify({'error': f'Failed to register column: {str(e)}'}), 500
        finally:
            conn.close()

    @app.route('/api/admin/excel-columns/delete', methods=['POST', 'DELETE'])
    def admin_delete_excel_column():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        data = request.get_json() or {}
        target_type = data.get('target_type', 'expense')
        
        privilege = 'Expense Columns List'
        if target_type == 'emi':
            privilege = 'EMI Columns List'
        elif target_type == 'excel':
            privilege = 'Excel Import & Export Columns'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'delete'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
        column_key = data.get('column_key')
        target_type = data.get('target_type', 'expense')
        
        if not column_key:
            return jsonify({'error': 'Column key is required.'}), 400
            
        system_keys = {
            'expense': ('date', 'category', 'amount'),
            'emi': ('name', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type')
        }
        if column_key in system_keys.get(target_type, ()):
            return jsonify({'error': 'Cannot delete system required columns.'}), 400
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'DELETE FROM excel_columns WHERE column_key = ? AND target_type = ?',
            (column_key, target_type)
        )
        conn.commit()
        rows_affected = cursor.rowcount
        conn.close()
        if rows_affected > 0:
            return jsonify({'success': True, 'message': 'Column removed successfully.'})
        else:
            return jsonify({'error': 'Column not found.'}), 404

    @app.route('/api/admin/excel-columns/update-single', methods=['POST'])
    def admin_update_single_excel_column():
        if not is_logged_in():
            return jsonify({'error': 'Unauthorized'}), 401
        data = request.get_json() or {}
        column_key = data.get('column_key', '').strip()
        target_type = data.get('target_type', 'expense').strip().lower()
        
        privilege = 'Expense Columns List'
        if target_type == 'emi':
            privilege = 'EMI Columns List'
        elif target_type == 'excel':
            privilege = 'Excel Import & Export Columns'
            
        if not database.check_backend_privilege(session['user_id'], privilege, 'edit'):
            return jsonify({'error': f'Forbidden: Missing privilege {privilege}'}), 403
            
        column_label = data.get('column_label')
        display_order = data.get('display_order')
        is_enabled_import = data.get('is_enabled_import')
        is_enabled_export = data.get('is_enabled_export')
        
        if is_enabled_import is not None:
            is_enabled_import = int(is_enabled_import)
        if is_enabled_export is not None:
            is_enabled_export = int(is_enabled_export)
        if display_order is not None:
            display_order = int(display_order)
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE excel_columns 
                SET column_label = COALESCE(?, column_label),
                    display_order = COALESCE(?, display_order),
                    is_enabled_import = COALESCE(?, is_enabled_import),
                    is_enabled_export = COALESCE(?, is_enabled_export)
                WHERE column_key = ? AND target_type = ?
                """,
                (column_label, display_order, is_enabled_import, is_enabled_export, column_key, target_type)
            )
            conn.commit()
            return jsonify({'success': True, 'message': 'Column updated successfully.'})
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

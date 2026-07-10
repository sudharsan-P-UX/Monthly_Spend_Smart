import unittest
import os
import sqlite3
import json
from werkzeug.security import generate_password_hash

# Mock dotenv module to prevent loading .env during test execution
import sys
import types
dotenv_mock = types.ModuleType('dotenv')
dotenv_mock.load_dotenv = lambda *args, **kwargs: None
sys.modules['dotenv'] = dotenv_mock

# Set environment before imports
os.environ['DATABASE_PATH'] = 'test_expenses.db'
for k in ['DATABASE_URL', 'POSTGRES_URL', 'POSTGRES_HOST', 'POSTGRES_PORT', 'POSTGRES_USER', 'POSTGRES_PASSWORD', 'POSTGRES_DB']:
    if k in os.environ:
        del os.environ[k]

# Import local modules
import database
from app import app

class ExpenseTrackerTestCase(unittest.TestCase):
    def setUp(self):
        # Override database path for testing
        database.DB_PATH = 'test_expenses.db'
        database.connection.DB_PATH = 'test_expenses.db'
        database.VercelDb.DB_PATH = 'test_expenses.db'
        
        # Clean up database tables for a fresh test run
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        tables_to_delete = [
            "UserRole", "UserLoginDetails", "Refusers", "RefRoleAccess", "RefRole",
            "expenses", "emis", "otps", "categories", "bank_modes", "payment_types",
            "payment_categories", "settings", "excel_columns", "Refcurreny", "user_expense_controls"
        ]
        for tbl in tables_to_delete:
            try:
                cursor.execute(f"DELETE FROM {tbl}")
            except Exception:
                pass
                
        # Reset sequences if PostgreSQL
        is_pg = conn.__class__.__name__ == 'PostgresConnectionWrapper'
        if is_pg:
            seeded_tables = [
                ('reffieldtype', 'fieldtypeid'),
                ('refrole', 'roleid'),
                ('refimportexport', 'importexportid'),
                ('refcurreny', 'currencyid'),
                ('categories', 'id'),
                ('bank_modes', 'id'),
                ('payment_types', 'id'),
                ('payment_categories', 'id'),
                ('refusers', 'loginid'),
                ('expenses', 'id'),
                ('emis', 'id')
            ]
            for table, pk in seeded_tables:
                try:
                    cursor.execute(f"SELECT setval(pg_get_serial_sequence('{table}', '{pk}'), 1, false)")
                except Exception:
                    pass
        try:
            conn.commit()
        except Exception:
            pass
        finally:
            conn.close()
            
        # Initialize test database (seeds everything fresh)
        database.init_db()
        
        # Flask testing client setup
        self.app = app.test_client()
        self.app.testing = True
        
    def tearDown(self):
        # Remove test database
        if os.path.exists('test_expenses.db'):
            try:
                os.remove('test_expenses.db')
            except OSError:
                pass

    def test_database_creation(self):
        """Test database and table creations"""
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        is_pg = conn.__class__.__name__ == 'PostgresConnectionWrapper'
        
        # Verify Refusers table exists
        if is_pg:
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_name = 'refusers'")
        else:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND LOWER(name)='refusers'")
        self.assertIsNotNone(cursor.fetchone())
        
        # Verify expenses table exists
        if is_pg:
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_name = 'expenses'")
        else:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND LOWER(name)='expenses'")
        self.assertIsNotNone(cursor.fetchone())
        
        # Verify new columns exist in expenses
        if is_pg:
            cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'expenses'")
            cols = [row[0].lower() for row in cursor.fetchall()]
        else:
            cursor.execute("PRAGMA table_info(expenses)")
            cols = [row['name'] for row in cursor.fetchall()]
            
        self.assertIn('bank_mode', cols)
        self.assertIn('payment_type', cols)
        self.assertIn('payment_category', cols)
        self.assertIn('payment_method', cols)
        
        conn.close()

    def test_user_creation_and_auth(self):
        """Test database user CRUD and checking pass hashes"""
        user_id = database.create_user('testuser', 'testpassword')
        self.assertIsNotNone(user_id)
        
        # Try duplicating user registration (should fail)
        dup_id = database.create_user('testuser', 'testpassword')
        self.assertIsNone(dup_id)
        
        # Retrieve user
        user = database.get_user_by_username('testuser')
        self.assertIsNotNone(user)
        self.assertEqual(user['username'], 'testuser')

    def test_logged_in_user_profile_and_password_endpoints(self):
        """Test fetching/updating user profile details and password as a logged in user"""
        # Create test user
        user_id = database.create_user('profileuser', 'pass123')
        
        # Log in
        with self.app.session_transaction() as sess:
            sess['user_id'] = user_id
            sess['username'] = 'profileuser'
            
        # Test GET profile details
        r_get = self.app.get('/api/user/profile')
        self.assertEqual(r_get.status_code, 200)
        data = json.loads(r_get.data.decode())
        self.assertEqual(data['username'], 'profileuser')
        self.assertEqual(data['first_name'], '')
        
        # Test POST update profile
        r_post = self.app.post('/api/user/profile/update', json={
            'first_name': 'John',
            'last_name': 'Doe',
            'email': 'john@example.com',
            'phone': '1234567891'
        })
        self.assertEqual(r_post.status_code, 200)
        res = json.loads(r_post.data.decode())
        self.assertTrue(res['success'])
        
        # Verify updated info in GET profile
        r_get_new = self.app.get('/api/user/profile')
        data_new = json.loads(r_get_new.data.decode())
        self.assertEqual(data_new['first_name'], 'John')
        self.assertEqual(data_new['last_name'], 'Doe')
        self.assertEqual(data_new['email'], 'john@example.com')
        self.assertEqual(data_new['phone'], '1234567891')
        
        # Test POST update password
        r_pass = self.app.post('/api/user/change-password', json={
            'new_password': 'newpass123',
            'confirm_password': 'newpass123'
        })
        self.assertEqual(r_pass.status_code, 200)
        res_pass = json.loads(r_pass.data.decode())
        self.assertTrue(res_pass['success'])

    def test_expense_crud(self):
        """Test adding, listing, updating, deleting expenses in DB"""
        user_id = database.create_user('expuser', 'password123')
        
        # Add expense with new fields
        exp_id = database.add_expense(user_id, 45.50, 'Food & Dining', 'Dinner with team', '2036-06-25', bank_mode='SBI', payment_type='GPay', payment_category='Salary', payment_method='Debit', status='Unpaid')
        self.assertIsNotNone(exp_id)
        
        # Get expense
        exp = database.get_expense_by_id(exp_id, user_id)
        self.assertIsNotNone(exp)
        self.assertEqual(exp['amount'], 45.50)
        self.assertEqual(exp['category'], 'Food & Dining')
        self.assertEqual(exp['bank_mode'], 'SBI')
        self.assertEqual(exp['payment_type'], 'GPay')
        self.assertEqual(exp['payment_category'], 'Salary')
        self.assertEqual(exp['payment_method'], 'Debit')
        self.assertEqual(exp['status'], 'Unpaid')
        
        # Filter check
        expenses = database.get_expenses(user_id, bank_mode='SBI')
        self.assertEqual(len(expenses), 1)
        
        # Filter check with status
        expenses_unpaid = database.get_expenses(user_id, status='Unpaid')
        self.assertEqual(len(expenses_unpaid), 1)
        
        # Filter paid expenses (should be 0, as status is Unpaid)
        expenses_paid = database.get_expenses(user_id, status='Paid')
        self.assertEqual(len(expenses_paid), 0)
        
        expenses_wrong = database.get_expenses(user_id, bank_mode='ICICI')
        self.assertEqual(len(expenses_wrong), 0)
        
        # Search check for description
        expenses_search = database.get_expenses(user_id, search='team')
        self.assertEqual(len(expenses_search), 1)
        
        # Search check for amount (partial match)
        expenses_amount_search = database.get_expenses(user_id, search='45')
        self.assertEqual(len(expenses_amount_search), 1)
        
        # Search check for amount (no match)
        expenses_amount_search_fail = database.get_expenses(user_id, search='999')
        self.assertEqual(len(expenses_amount_search_fail), 0)
        
        # Update expense
        success = database.update_expense(exp_id, user_id, 50.00, 'Food & Dining', 'Dinner with team (updated)', '2036-06-25', bank_mode='Kotak', payment_type='PhonePe', payment_category='Loan', interest=5.00, payment_method='Credit', status='Paid')
        self.assertTrue(success)
        
        updated_exp = database.get_expense_by_id(exp_id, user_id)
        self.assertEqual(updated_exp['amount'], 50.00)
        self.assertEqual(updated_exp['description'], 'Dinner with team (updated)')
        self.assertEqual(updated_exp['bank_mode'], 'Kotak')
        self.assertEqual(updated_exp['payment_type'], 'PhonePe')
        self.assertEqual(updated_exp['payment_category'], 'Loan')
        self.assertEqual(updated_exp['payment_method'], 'Credit')
        self.assertEqual(updated_exp['status'], 'Paid')
        
        # Delete expense
        del_success = database.delete_expense(exp_id, user_id)
        self.assertTrue(del_success)
        
        deleted_exp = database.get_expense_by_id(exp_id, user_id)
        self.assertIsNone(deleted_exp)

    def test_api_endpoints_without_login(self):
        """Verify API guards return 401 unauthorized when not logged in"""
        resp = self.app.get('/api/expenses')
        self.assertEqual(resp.status_code, 401)
        
        resp = self.app.post('/api/expenses/add', data=json.dumps({}), content_type='application/json')
        self.assertEqual(resp.status_code, 401)
        
        resp = self.app.get('/api/overview')
        self.assertEqual(resp.status_code, 401)

    def test_flask_auth_endpoints(self):
        """Test API endpoints for registration and login"""
        # 1. Register via API
        resp = self.app.post('/register', data=json.dumps({
            'username': 'flaskuser',
            'password': 'securepassword'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])
        
        # 2. Log out so we are not logged in during duplicate registration test
        self.app.get('/logout')

        # 3. Duplicate registration fails (username exists in database)
        resp = self.app.post('/register', data=json.dumps({
            'username': 'flaskuser',
            'password': 'securepassword'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 409)
        
        # 4. Log out again just in case
        self.app.get('/logout')
        
        # 4. Login via API
        resp = self.app.post('/login', data=json.dumps({
            'username': 'flaskuser',
            'password': 'securepassword'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])
        # 5. Check logged in status
        resp = self.app.get('/api/user/privileges')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(data['username'], 'flaskuser')

    def test_rbac_database_seeding(self):
        """Verify default roles, privileges, and admin seeding"""
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Verify roles
        roles = cursor.execute("SELECT RoleId as id, RoleName as name FROM RefRole ORDER BY RoleId").fetchall()
        self.assertEqual(len(roles), 3)
        self.assertEqual(roles[0]['name'], 'Admin')
        self.assertEqual(roles[1]['name'], 'User')
        self.assertEqual(roles[2]['name'], 'Viewer')
        
        # Verify privileges for Viewer
        priv_row = cursor.execute("SELECT Editaccess, DeleteAccess, Addaccess FROM RefRoleAccess WHERE RoleId = 3").fetchone()
        self.assertIsNotNone(priv_row)
        
        def parse_bit(val):
            if val is None:
                return 1
            if isinstance(val, str):
                return 1 if '1' in val else 0
            if isinstance(val, bytes):
                return 1 if b'\x01' in val or b'1' in val else 0
            return 1 if bool(val) else 0
            
        viewer_privs = {
            'can_view': 1,
            'can_add': parse_bit(priv_row[2]),
            'can_edit': parse_bit(priv_row[0]),
            'can_delete': parse_bit(priv_row[1])
        }
        
        self.assertEqual(viewer_privs['can_view'], 1)
        self.assertEqual(viewer_privs['can_add'], 0)
        self.assertEqual(viewer_privs['can_edit'], 0)
        self.assertEqual(viewer_privs['can_delete'], 0)
        
        # Verify admin user exists
        admin = cursor.execute("SELECT * FROM Refusers WHERE Username = 'admin'").fetchone()
        self.assertIsNotNone(admin)
        
        admin_id = admin['loginid'] if 'loginid' in admin else admin[0]
        user_role = cursor.execute("SELECT RoleId FROM UserRole WHERE LoginId = ?", (admin_id,)).fetchone()
        self.assertIsNotNone(user_role)
        role_id = user_role['roleid'] if 'roleid' in user_role else user_role[0]
        self.assertEqual(role_id, 1)
        
        # Verify default categories are seeded
        cats_count = cursor.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        self.assertGreaterEqual(cats_count, 9)
        
        conn.close()

    def test_rbac_api_protection(self):
        """Test API endpoint restriction for users with Viewer role"""
        # Register a viewer user
        # Note: default created user role is User (2). We will manually update it to Viewer (3).
        user_id = database.create_user('vieweruser', 'password123', role_id=3)
        self.assertIsNotNone(user_id)
        
        # Log in as vieweruser
        self.app.post('/login', data=json.dumps({
            'username': 'vieweruser',
            'password': 'password123'
        }), content_type='application/json')
        
        # 1. View expenses should succeed (starts empty, returns 200)
        resp = self.app.get('/api/expenses')
        self.assertEqual(resp.status_code, 200)
        
        # 2. Add expense should return 403 Forbidden
        resp = self.app.post('/api/expenses/add', data=json.dumps({
            'amount': 100.0,
            'category': 'Other',
            'date': '2026-06-25'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 403)
        
        # 3. Edit expense should return 403 Forbidden
        resp = self.app.post('/api/expenses/edit/1', data=json.dumps({
            'amount': 150.0,
            'category': 'Other',
            'date': '2026-06-25'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 403)
        
        # 4. Delete expense should return 403 Forbidden
        resp = self.app.post('/api/expenses/delete/1', data=json.dumps({}), content_type='application/json')
        self.assertEqual(resp.status_code, 403)
        
        # 5. Access admin panel endpoints should return 403 Forbidden
        resp = self.app.get('/api/admin/users')
        self.assertEqual(resp.status_code, 403)

    def test_categories_crud_and_sorting(self):
        """Test category creation, ordering, and deletion as Admin"""
        # Log in as default admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)
        
        # 1. Create a category with order -50 (should come first)
        resp = self.app.post('/api/admin/categories/create', data=json.dumps({
            'name': 'AAA_FirstCat',
            'display_order': -50
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        cat_id = data['id']
        
        # 2. Get categories and verify AAA_FirstCat is indeed first
        resp = self.app.get('/api/categories')
        self.assertEqual(resp.status_code, 200)
        cats = json.loads(resp.data)
        self.assertEqual(cats[0]['name'], 'AAA_FirstCat')
        
        # 3. Update AAA_FirstCat order to 500 (should move down the list)
        resp = self.app.post(f'/api/admin/categories/edit/{cat_id}', data=json.dumps({
            'name': 'AAA_FirstCat_Updated',
            'display_order': 500
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # 4. Verify order changes
        resp = self.app.get('/api/categories')
        cats = json.loads(resp.data)
        self.assertNotEqual(cats[0]['name'], 'AAA_FirstCat_Updated')
        self.assertEqual(cats[-1]['name'], 'AAA_FirstCat_Updated')
        
        # 5. Delete category
        resp = self.app.delete(f'/api/admin/categories/delete/{cat_id}')
        self.assertEqual(resp.status_code, 200)
        
        # Verify it is deleted
        resp = self.app.get('/api/categories')
        cats = json.loads(resp.data)
        cat_names = [c['name'] for c in cats]
        self.assertNotIn('AAA_FirstCat_Updated', cat_names)

    def test_dropdowns_crud_and_sorting(self):
        """Test Bank Modes, Payment Types, and Payment Categories CRUD as Admin"""
        # Log in as default admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        # ---- BANK MODES ----
        # 1. Create a bank mode
        resp = self.app.post('/api/admin/bank_modes/create', data=json.dumps({
            'name': 'AAA_TestBank',
            'display_order': -10
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        bm_id = data['id']

        # 2. Get bank modes and verify order
        resp = self.app.get('/api/bank_modes')
        self.assertEqual(resp.status_code, 200)
        bms = json.loads(resp.data)
        self.assertEqual(bms[0]['name'], 'AAA_TestBank')

        # 3. Edit bank mode
        resp = self.app.post(f'/api/admin/bank_modes/edit/{bm_id}', data=json.dumps({
            'name': 'AAA_TestBank_Updated',
            'display_order': 200
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify change
        resp = self.app.get('/api/bank_modes')
        bms = json.loads(resp.data)
        self.assertEqual(bms[-1]['name'], 'AAA_TestBank_Updated')

        # 4. Delete bank mode
        resp = self.app.delete(f'/api/admin/bank_modes/delete/{bm_id}')
        self.assertEqual(resp.status_code, 200)

        # Verify deleted
        resp = self.app.get('/api/bank_modes')
        bm_names = [b['name'] for b in json.loads(resp.data)]
        self.assertNotIn('AAA_TestBank_Updated', bm_names)

        # ---- PAYMENT TYPES ----
        # 1. Create a payment type
        resp = self.app.post('/api/admin/payment_types/create', data=json.dumps({
            'name': 'AAA_TestType',
            'display_order': -10
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        pt_id = data['id']

        # 2. Get payment types and verify order
        resp = self.app.get('/api/payment_types')
        self.assertEqual(resp.status_code, 200)
        pts = json.loads(resp.data)
        self.assertEqual(pts[0]['name'], 'AAA_TestType')

        # 3. Edit payment type
        resp = self.app.post(f'/api/admin/payment_types/edit/{pt_id}', data=json.dumps({
            'name': 'AAA_TestType_Updated',
            'display_order': 200
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify change
        resp = self.app.get('/api/payment_types')
        pts = json.loads(resp.data)
        self.assertEqual(pts[-1]['name'], 'AAA_TestType_Updated')

        # 4. Delete payment type
        resp = self.app.delete(f'/api/admin/payment_types/delete/{pt_id}')
        self.assertEqual(resp.status_code, 200)

        # Verify deleted
        resp = self.app.get('/api/payment_types')
        pt_names = [p['name'] for p in json.loads(resp.data)]
        self.assertNotIn('AAA_TestType_Updated', pt_names)

        # ---- PAYMENT CATEGORIES ----
        # 1. Create a payment category
        resp = self.app.post('/api/admin/payment_categories/create', data=json.dumps({
            'name': 'AAA_TestPayCat',
            'display_order': -10
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        pc_id = data['id']

        # 2. Get payment categories and verify order
        resp = self.app.get('/api/payment_categories')
        self.assertEqual(resp.status_code, 200)
        pcs = json.loads(resp.data)
        self.assertEqual(pcs[0]['name'], 'AAA_TestPayCat')

        # 3. Edit payment category
        resp = self.app.post(f'/api/admin/payment_categories/edit/{pc_id}', data=json.dumps({
            'name': 'AAA_TestPayCat_Updated',
            'display_order': 200
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify change
        resp = self.app.get('/api/payment_categories')
        pcs = json.loads(resp.data)
        self.assertEqual(pcs[-1]['name'], 'AAA_TestPayCat_Updated')

        # 4. Delete payment category
        resp = self.app.delete(f'/api/admin/payment_categories/delete/{pc_id}')
        self.assertEqual(resp.status_code, 200)

        # Verify deleted
        resp = self.app.get('/api/payment_categories')
        pc_names = [p['name'] for p in json.loads(resp.data)]
        self.assertNotIn('AAA_TestPayCat_Updated', pc_names)

    def test_role_renaming_and_guards(self):
        """Test renaming roles and security constraints"""
        # Log in as default admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        # 1. Rename role ID 2 (User)
        resp = self.app.post('/api/admin/roles/edit', data=json.dumps({
            'role_id': 2,
            'name': 'Standard User'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify it got updated
        resp = self.app.get('/api/admin/roles')
        roles = json.loads(resp.data)
        user_role = next(r for r in roles if r['id'] == 2)
        self.assertEqual(user_role['name'], 'Standard User')

        # 2. Try to rename Admin role (role_id=1), which must fail
        resp = self.app.post('/api/admin/roles/edit', data=json.dumps({
            'role_id': 1,
            'name': 'Super Admin'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 400)

        # Verify Admin role name did not change
        resp = self.app.get('/api/admin/roles')
        roles = json.loads(resp.data)
        admin_role = next(r for r in roles if r['id'] == 1)
        self.assertEqual(admin_role['name'], 'Admin')

    def test_month_year_filters_and_year_totals(self):
        """Test month, year filtering, year totals API, and interest tracking"""
        # Register and log in a user
        user_id = database.create_user('filteruser', 'password123', role_id=2)
        self.assertIsNotNone(user_id)
        
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'filteruser',
            'password': 'password123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)
        
        # Add various expenses across different months/years/modes
        # 1. 2026-06-25, Food & Dining, $100.0, Debit
        database.add_expense(user_id, 100.0, 'Food & Dining', 'Expense 1', '2026-06-25', bank_mode='SBI', payment_type='GPay', payment_category='Salary', payment_method='Debit')
        # 2. 2026-06-10, Shopping, $50.0, Credit, Interest: $5.0
        database.add_expense(user_id, 50.0, 'Shopping', 'Expense 2', '2026-06-10', bank_mode='Kotak', payment_type='PhonePe', payment_category='Credit Card', payment_method='Credit', interest=5.0)
        # 3. 2026-07-01, Utilities, $200.0, Savings (which should count as Debit)
        database.add_expense(user_id, 200.0, 'Utilities', 'Expense 3', '2026-07-01', bank_mode='ICICI', payment_type='BHIM', payment_category='Loan', payment_method='Debit')
        # 4. 2025-12-15, Housing & Rent, $500.0, Salary (which should count as Credit)
        database.add_expense(user_id, 500.0, 'Housing & Rent', 'Expense 4', '2025-12-15', bank_mode='SBI', payment_type='Cash', payment_category='Borrow', payment_method='Credit', interest=12.5)
        
        # Test /api/expenses filters
        # Filter by Month 06
        resp = self.app.get('/api/expenses?month=06')
        self.assertEqual(resp.status_code, 200)
        expenses_06 = json.loads(resp.data)
        self.assertEqual(len(expenses_06), 2)
        
        # Filter by Year 2026
        resp = self.app.get('/api/expenses?year=2026')
        self.assertEqual(resp.status_code, 200)
        expenses_2026 = json.loads(resp.data)
        self.assertEqual(len(expenses_2026), 3)
        
        # Test /api/year_totals
        # Year 2026 totals: Debit should be 100 (Debit) + 200 (Savings) = 300; Credit should be 50 (Credit)
        resp = self.app.get('/api/year_totals?year=2026')
        self.assertEqual(resp.status_code, 200)
        totals_2026 = json.loads(resp.data)
        self.assertEqual(totals_2026['debit'], 300.0)
        self.assertEqual(totals_2026['credit'], 50.0)
        
        # Year 2025 totals: Debit should be 0; Credit should be 500 (Salary is Credit)
        resp = self.app.get('/api/year_totals?year=2025')
        self.assertEqual(resp.status_code, 200)
        totals_2025 = json.loads(resp.data)
        self.assertEqual(totals_2025['debit'], 0.0)
        self.assertEqual(totals_2025['credit'], 500.0)
        
        # Test /api/overview filters
        # Filtered to 2026-06
        resp = self.app.get('/api/overview?month=06&year=2026')
        self.assertEqual(resp.status_code, 200)
        overview_06_2026 = json.loads(resp.data)
        self.assertEqual(overview_06_2026['month_debit'], 100.0)
        self.assertEqual(overview_06_2026['month_credit'], 50.0)
        self.assertEqual(overview_06_2026['month_interest'], 5.0)
        self.assertEqual(overview_06_2026['total_interest'], 17.5)

        # Test adding expense via API with interest
        resp = self.app.post('/api/expenses/add', data=json.dumps({
            'amount': 300.0,
            'category': 'Interest Gain',
            'date': '2026-06-15',
            'payment_category': 'Credit Card',
            'payment_method': 'Credit',
            'interest': 15.0
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        add_data = json.loads(resp.data)
        self.assertTrue(add_data['success'])
        exp_id = add_data['id']

        # Get exp and check interest
        exp_obj = database.get_expense_by_id(exp_id, user_id)
        self.assertEqual(exp_obj['interest'], 15.0)

        # Edit expense via API
        resp = self.app.post(f'/api/expenses/edit/{exp_id}', data=json.dumps({
            'amount': 350.0,
            'category': 'Interest Gain Updated',
            'date': '2026-06-15',
            'payment_category': 'Credit Card',
            'payment_method': 'Credit',
            'interest': 22.0
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        exp_obj_updated = database.get_expense_by_id(exp_id, user_id)
        self.assertEqual(exp_obj_updated['interest'], 22.0)

    def test_excel_import_export(self):
        """Test excel import, export, and template download endpoints"""
        from io import BytesIO
        from openpyxl import Workbook
        import datetime

        # 1. Register and log in a user
        user_id = database.create_user('exceluser', 'password123', role_id=2)
        self.assertIsNotNone(user_id)
        
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'exceluser',
            'password': 'password123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        # 2. Test template download
        template_resp = self.app.get('/api/expenses/import-template')
        self.assertEqual(template_resp.status_code, 200)
        self.assertEqual(template_resp.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 3. Test export (empty first)
        export_resp = self.app.get('/api/expenses/export')
        self.assertEqual(export_resp.status_code, 200)
        self.assertEqual(export_resp.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Add an expense and test export
        database.add_expense(user_id, 80.0, 'Shopping', 'Export test', '2026-06-25', bank_mode='Kotak', payment_type='PhonePe', payment_category='Credit Card', payment_method='Credit', interest=5.0)
        
        export_resp = self.app.get('/api/expenses/export?category=Shopping')
        self.assertEqual(export_resp.status_code, 200)
        self.assertEqual(export_resp.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 4. Test import endpoint
        wb = Workbook()
        ws = wb.active
        ws.title = "Import"
        
        # Headers
        headers = ["Date", "Category", "Description", "Gateway", "Bank", "Source", "Method", "Amount", "Interest"]
        ws.append(headers)
        
        # Row 1: Valid Debit, interest specified but should be overridden to 0
        ws.append(["2026-06-25", "Food", "Lunch", "GPay", "SBI", "Salary", "Debit", 15.50, 2.00])
        # Row 2: Valid Credit, datetime date, valid interest
        ws.append([datetime.date(2026, 6, 26), "Shopping", "Clothes", "Credit Card", "Kotak", "Credit Card", "Credit", 120.00, 10.00])
        # Row 3: Invalid amount (string instead of float, should be skipped)
        ws.append(["2026-06-27", "Utilities", "Power", "Netbanking", "ICICI", "Savings", "Debit", "abc", 0.00])
        # Row 4: Invalid date format (should be skipped)
        ws.append(["invalid-date", "Entertainment", "Movie", "Cash", "None", "None", "Debit", 20.00, 0.00])
        # Row 5: Missing required column - Date (should be skipped)
        ws.append([None, "Medical", "Checkup", "GPay", "SBI", "Salary", "Debit", 50.00, 0.00])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Post file
        import_resp = self.app.post('/api/expenses/import', data={
            'file': (excel_file, 'test_import.xlsx')
        }, content_type='multipart/form-data')
        
        self.assertEqual(import_resp.status_code, 200)
        import_data = json.loads(import_resp.data)
        self.assertTrue(import_data['success'])
        self.assertEqual(import_data['imported'], 2)  # Rows 1 and 2
        self.assertEqual(import_data['skipped'], 3)   # Rows 3, 4, and 5

        # Check DB entries for exceluser
        expenses = database.get_expenses(user_id)
        # Note: we added 1 manually before, plus 2 imported = 3 total
        self.assertEqual(len(expenses), 3)

        # Verify Row 1 details (Debit, interest overridden to 0)
        lunch_exp = next(e for e in expenses if e['description'] == 'Lunch')
        self.assertEqual(lunch_exp['amount'], 15.50)
        self.assertEqual(lunch_exp['payment_method'], 'Debit')
        self.assertEqual(lunch_exp['interest'], 0.0)

        # Verify Row 2 details (Credit, interest remains 10)
        clothes_exp = next(e for e in expenses if e['description'] == 'Clothes')
        self.assertEqual(clothes_exp['amount'], 120.00)
        self.assertEqual(clothes_exp['payment_method'], 'Credit')
        self.assertEqual(clothes_exp['interest'], 10.0)

    def test_excel_columns_admin_configuration(self):
        """Test excel columns admin toggle functionality and its effect on import/export"""
        from io import BytesIO
        from openpyxl import load_workbook, Workbook
        import datetime

        # 1. Login as admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        # 2. Get excel columns list
        resp = self.app.get('/api/admin/excel-columns')
        self.assertEqual(resp.status_code, 200)
        cols = json.loads(resp.data)
        self.assertEqual(len(cols), 10)

        # Find "interest" and verify it's enabled by default
        interest_col = next(c for c in cols if c['column_key'] == 'interest')
        self.assertEqual(interest_col['is_enabled_import'], 1)
        self.assertEqual(interest_col['is_enabled_export'], 1)
        self.assertEqual(interest_col['is_required'], 0)

        # 3. Toggle "interest" column to disabled for import & export
        resp = self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'interest',
            'is_enabled': 0,
            'type_key': 'import'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(json.loads(resp.data)['success'])

        resp = self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'interest',
            'is_enabled': 0,
            'type_key': 'export'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(json.loads(resp.data)['success'])

        # Verify it was updated
        resp = self.app.get('/api/admin/excel-columns')
        cols = json.loads(resp.data)
        interest_col = next(c for c in cols if c['column_key'] == 'interest')
        self.assertEqual(interest_col['is_enabled_import'], 0)
        self.assertEqual(interest_col['is_enabled_export'], 0)

        # 4. Disable a required column for admin (should succeed since admin has "all access")
        resp = self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'date',
            'is_enabled': 0,
            'type_key': 'import'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(json.loads(resp.data)['success'])

        # Verify date is disabled
        resp = self.app.get('/api/admin/excel-columns')
        cols = json.loads(resp.data)
        date_col = next(c for c in cols if c['column_key'] == 'date')
        self.assertEqual(date_col['is_enabled_import'], 0)

        # 5. Check import template (interest and date should not be in the headers)
        template_resp = self.app.get('/api/expenses/import-template')
        self.assertEqual(template_resp.status_code, 200)
        wb = load_workbook(BytesIO(template_resp.data), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn('Interest', headers)
        self.assertNotIn('Date', headers)
        self.assertIn('Amount', headers)

        # 6. Check export (interest should not be in the headers)
        export_resp = self.app.get('/api/expenses/export')
        self.assertEqual(export_resp.status_code, 200)
        wb_exp = load_workbook(BytesIO(export_resp.data), data_only=True)
        ws_exp = wb_exp.active
        headers_exp = [cell.value for cell in ws_exp[1]]
        self.assertNotIn('Interest', headers_exp)

        # 7. Check import with disabled column (and missing required date since date is disabled):
        # If we upload a file, it will ignore interest and fallback to default date (today)
        wb_imp = Workbook()
        ws_imp = wb_imp.active
        ws_imp.append(["Category", "Amount"]) # Date and Interest are omitted
        ws_imp.append(["Food", 15.50])
        excel_file = BytesIO()
        wb_imp.save(excel_file)
        excel_file.seek(0)

        import_resp = self.app.post('/api/expenses/import', data={
            'file': (excel_file, 'test_import.xlsx')
        }, content_type='multipart/form-data')
        self.assertEqual(import_resp.status_code, 200)
        
        # Verify expense is created with default today's date and interest is 0.0
        expenses = database.get_expenses(1)
        food_exp = next(e for e in expenses if e['category'] == 'Food' and e['amount'] == 15.50)
        self.assertEqual(food_exp['interest'], 0.0)
        self.assertEqual(food_exp['date'], datetime.date.today().strftime('%Y-%m-%d'))

        # 8. Test read-only access for non-admin
        self.app.get('/logout')
        # Register a non-admin user
        reg_resp = self.app.post('/register', data={
            'username': 'normaluser',
            'password': 'normaluser123'
        })
        self.assertEqual(reg_resp.status_code, 200)

        # Non-admin gets excel-columns list successfully (returns 200)
        resp = self.app.get('/api/admin/excel-columns')
        self.assertEqual(resp.status_code, 200)

        # Non-admin tries to toggle columns (should fail with 403 Forbidden)
        resp = self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'interest',
            'is_enabled': 1,
            'type_key': 'import'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 403)

        # 9. Restore interest and date status for other tests
        self.app.get('/logout')
        # Login back as admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'interest',
            'is_enabled': 1,
            'type_key': 'import'
        }), content_type='application/json')
        self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'interest',
            'is_enabled': 1,
            'type_key': 'export'
        }), content_type='application/json')
        self.app.post('/api/admin/excel-columns/toggle', data=json.dumps({
            'column_key': 'date',
            'is_enabled': 1,
            'type_key': 'import'
        }), content_type='application/json')

    def test_change_password_flow(self):
        """Test username verification and password change flows"""
        # 1. Verify non-existent username
        resp = self.app.get('/api/verify-username?username=doesnotexist')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertFalse(data['exists'])

        # 2. Verify existing admin username
        resp = self.app.get('/api/verify-username?username=admin')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['exists'])

        # 3. Change password for non-existent username
        resp = self.app.post('/api/change-password', data=json.dumps({
            'username': 'doesnotexist',
            'new_password': 'newpassword123',
            'confirm_password': 'newpassword123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 404)

        # 4. Change password with mismatched passwords
        resp = self.app.post('/api/change-password', data=json.dumps({
            'username': 'admin',
            'new_password': 'newpassword123',
            'confirm_password': 'mismatched123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 400)

        # 5. Change password with short password
        resp = self.app.post('/api/change-password', data=json.dumps({
            'username': 'admin',
            'new_password': '123',
            'confirm_password': '123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 400)

        # 6. Change password successfully
        resp = self.app.post('/api/change-password', data=json.dumps({
            'username': 'admin',
            'new_password': 'newadmin12345',
            'confirm_password': 'newadmin12345'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # 7. Try login with old password (should fail)
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 401)

        # 8. Try login with new password (should succeed)
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'newadmin12345'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Log out
        self.app.get('/logout')

        # 9. Create a test user, log in as admin, and change that user's password via Admin API
        test_user_id = database.create_user('testpwduser', 'user123', role_id=2)
        self.assertIsNotNone(test_user_id)

        # Try admin change password without admin login (should return 401)
        resp = self.app.post('/api/admin/users/change_password', data=json.dumps({
            'user_id': test_user_id,
            'new_password': 'newuser123',
            'confirm_password': 'newuser123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 401)

        # Log in as admin
        self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'newadmin12345'
        }), content_type='application/json')

        # Change password via Admin API
        resp = self.app.post('/api/admin/users/change_password', data=json.dumps({
            'user_id': test_user_id,
            'new_password': 'newuser123',
            'confirm_password': 'newuser123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # Log out admin
        self.app.get('/logout')

        # Verify new user password works
        resp = self.app.post('/login', data=json.dumps({
            'username': 'testpwduser',
            'password': 'newuser123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

    def test_emi_flow(self):
        """Test EMI CRUD operations for both normal users and administrators"""
        # 1. Fetch EMIs without login
        resp = self.app.get('/api/emis')
        self.assertEqual(resp.status_code, 401)

        # 2. Register/Login a user
        self.app.post('/register', data=json.dumps({
            'username': 'emiuser',
            'password': 'password123'
        }), content_type='application/json')

        # 3. Fetch EMIs (should be empty initially)
        resp = self.app.get('/api/emis')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(len(data), 0)

        # 4. Add an EMI
        resp = self.app.post('/api/emis/add', data=json.dumps({
            'name': 'Car Loan',
            'principal_amount': 20000.0,
            'emi_amount': 500.0,
            'start_date': '2026-07-01',
            'end_date': '2029-07-01',
            'tenure_months': 36,
            'interest_rate': 8.5,
            'due_date': '5',
            'payment_type': 'Auto',
            'payment_gateway': 'GPay',
            'payment_bank': 'SBI'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])
        emi_id = data['id']

        # 5. Add an EMI with missing fields (should fail)
        resp = self.app.post('/api/emis/add', data=json.dumps({
            'name': 'Invalid Loan'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 400)

        # 6. Fetch EMIs (should return 1 EMI)
        resp = self.app.get('/api/emis')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['name'], 'Car Loan')
        self.assertEqual(data[0]['emi_amount'], 500.0)

        # 7. Edit the EMI
        resp = self.app.post(f'/api/emis/edit/{emi_id}', data=json.dumps({
            'name': 'Car Loan (Updated)',
            'principal_amount': 20000.0,
            'emi_amount': 450.0,
            'start_date': '2026-07-01',
            'end_date': '2029-07-01',
            'tenure_months': 36,
            'interest_rate': 8.5,
            'due_date': '10',
            'payment_type': 'Manual',
            'payment_gateway': 'PhonePe',
            'payment_bank': 'Kotak'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify edits
        resp = self.app.get('/api/emis')
        data = json.loads(resp.data)
        self.assertEqual(data[0]['name'], 'Car Loan (Updated)')
        self.assertEqual(data[0]['emi_amount'], 450.0)
        self.assertEqual(data[0]['due_date'], '10')
        self.assertEqual(data[0]['payment_type'], 'Manual')

        # Test user EMI export
        resp = self.app.get('/api/emis/export')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Test user EMI import
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.append(["EMI Name", "Loan Amount", "Interest Rate", "Tenure", "Monthly EMI", "Start Date", "End Date", "Due Date", "Payment Type"])
        ws.append(["Imported User Loan", 10000.0, 5.0, 10, 1050.0, "2026-07-01", "2027-05-01", "5", "Auto"])
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        resp = self.app.post('/api/emis/import', data={
            'file': (out, 'test.xlsx')
        }, content_type='multipart/form-data')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # Log out user
        self.app.get('/logout')

        # 8. Log in as Admin
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # 9. Admin fetch EMIs
        resp = self.app.get('/api/admin/emis')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertGreaterEqual(len(data), 1)
        # Verify user association in admin view
        user_emi = next(e for e in data if e['id'] == emi_id)
        self.assertEqual(user_emi['username'], 'emiuser')

        # 10. Admin create EMI for user emiuser
        user = database.get_user_by_username('emiuser')
        user_id = user['id']
        resp = self.app.post('/api/admin/emis/create', data=json.dumps({
            'user_id': user_id,
            'name': 'Admin Created Loan',
            'principal_amount': 5000.0,
            'emi_amount': 150.0,
            'start_date': '2026-08-01',
            'end_date': '2027-08-01',
            'tenure_months': 12,
            'interest_rate': 10.0,
            'due_date': '1',
            'payment_type': 'Auto',
            'payment_gateway': 'Paytm',
            'payment_bank': 'ICICI'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        admin_emi_id = data['id']

        # 11. Admin edit EMI
        resp = self.app.post(f'/api/admin/emis/edit/{admin_emi_id}', data=json.dumps({
            'name': 'Admin Created Loan (Updated)',
            'principal_amount': 5000.0,
            'emi_amount': 160.0,
            'start_date': '2026-08-01',
            'end_date': '2027-08-01',
            'tenure_months': 12,
            'interest_rate': 10.0,
            'due_date': '1',
            'payment_type': 'Auto',
            'payment_gateway': 'Paytm',
            'payment_bank': 'ICICI'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Test admin EMI export
        resp = self.app.get('/api/admin/emis/export')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Test admin EMI import
        wb_admin = Workbook()
        ws_admin = wb_admin.active
        ws_admin.append(["Username", "EMI Name", "Loan Amount", "Interest Rate", "Tenure", "Monthly EMI", "Start Date", "End Date", "Due Date", "Payment Type"])
        ws_admin.append(["emiuser", "Imported Admin Loan", 10000.0, 5.0, 10, 1050.0, "2026-07-01", "2027-05-01", "5", "Auto"])
        
        out_admin = BytesIO()
        wb_admin.save(out_admin)
        out_admin.seek(0)
        
        resp = self.app.post('/api/admin/emis/import', data={
            'file': (out_admin, 'test.xlsx')
        }, content_type='multipart/form-data')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # 12. Admin delete EMI
        resp = self.app.delete(f'/api/admin/emis/delete/{admin_emi_id}')
        self.assertEqual(resp.status_code, 200)

        # Log out admin
        self.app.get('/logout')

        # Log in back as user to delete their original EMI
        resp = self.app.post('/login', data=json.dumps({
            'username': 'emiuser',
            'password': 'password123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # 13. User delete EMI
        resp = self.app.delete(f'/api/emis/delete/{emi_id}')
        self.assertEqual(resp.status_code, 200)

        # 14. Fetch EMIs (should be empty after deleting remaining imported EMIs)
        resp = self.app.get('/api/emis')
        data = json.loads(resp.data)
        for e in data:
            self.app.delete(f'/api/emis/delete/{e["id"]}')
            
        resp = self.app.get('/api/emis')
        data = json.loads(resp.data)
        self.assertEqual(len(data), 0)

    def test_currency_flow(self):
        # 1. Fetch default active currency
        resp = self.app.get('/api/active-currency')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(data['symbol'], '₹')
        self.assertEqual(data['country'], 'India')

        # Log in as admin
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # 2. Get all currencies
        resp = self.app.get('/api/admin/currencies')
        self.assertEqual(resp.status_code, 200)
        currs = json.loads(resp.data)
        self.assertTrue(len(currs) >= 4)

        # 3. Add a new currency
        resp = self.app.post('/api/admin/currencies/add', data=json.dumps({
            'country': 'Japan',
            'country_desc': 'Japanese Yen',
            'symbol': '¥'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])
        new_curr_id = data['id']

        # 4. Set the new currency as active
        resp = self.app.post(f'/api/admin/currencies/set_active/{new_curr_id}')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # 5. Verify the active currency changed globally
        resp = self.app.get('/api/active-currency')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(data['symbol'], '¥')
        self.assertEqual(data['country'], 'Japan')

        # 6. Delete the currency configuration
        resp = self.app.delete(f'/api/admin/currencies/delete/{new_curr_id}')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # 7. Verify it falls back or reverts active currency
        resp = self.app.get('/api/active-currency')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertNotEqual(data['country'], 'Japan')

        # Log out admin
        self.app.get('/logout')

    def test_custom_excel_columns_and_inline_creation(self):
        """Test custom excel column creation, deletion, target filtering, and inline controls creation"""
        # 1. Login as admin
        login_resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(login_resp.status_code, 200)

        # 2. Add custom column for expense
        resp = self.app.post('/api/admin/excel-columns/create', data=json.dumps({
            'column_key': 'test_custom_col',
            'column_label': 'Test Custom Column',
            'target_type': 'expense',
            'is_enabled_import': 1,
            'is_enabled_export': 1,
            'is_required': 0
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertTrue(data['success'])

        # Check in get list by target_type
        resp = self.app.get('/api/admin/excel-columns?target_type=expense')
        self.assertEqual(resp.status_code, 200)
        cols = json.loads(resp.data)
        custom_col = next((c for c in cols if c['column_key'] == 'test_custom_col'), None)
        self.assertIsNotNone(custom_col)
        self.assertEqual(custom_col['column_label'], 'Test Custom Column')
        self.assertEqual(custom_col['target_type'], 'expense')

        # Check get list for EMI doesn't have it
        resp = self.app.get('/api/admin/excel-columns?target_type=emi')
        self.assertEqual(resp.status_code, 200)
        cols_emi = json.loads(resp.data)
        self.assertFalse(any(c['column_key'] == 'test_custom_col' for c in cols_emi))

        # 3. Create a custom column for EMI
        resp = self.app.post('/api/admin/excel-columns/create', data=json.dumps({
            'column_key': 'test_custom_emi_col',
            'column_label': 'Test Custom EMI Column',
            'target_type': 'emi',
            'is_enabled_import': 1,
            'is_enabled_export': 1,
            'is_required': 0
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Check it is in EMI columns
        resp = self.app.get('/api/admin/excel-columns?target_type=emi')
        cols_emi2 = json.loads(resp.data)
        self.assertTrue(any(c['column_key'] == 'test_custom_emi_col' for c in cols_emi2))

        # 4. Delete the custom columns
        resp = self.app.post('/api/admin/excel-columns/delete', data=json.dumps({
            'column_key': 'test_custom_col',
            'target_type': 'expense'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        resp = self.app.post('/api/admin/excel-columns/delete', data=json.dumps({
            'column_key': 'test_custom_emi_col',
            'target_type': 'emi'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Check they are deleted
        resp = self.app.get('/api/admin/excel-columns?target_type=expense')
        self.assertFalse(any(c['column_key'] == 'test_custom_col' for c in json.loads(resp.data)))

        # 5. Test unified control items creation inline
        resp = self.app.post('/api/admin/categories/create', data=json.dumps({
            'name': 'Test Inline Cat',
            'display_order': 5
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Verify it got added
        resp = self.app.get('/api/categories')
        cats = json.loads(resp.data)
        self.assertTrue(any(c['name'] == 'Test Inline Cat' for c in cats))

        # Log out admin
        self.app.get('/logout')

    def test_custom_column_parent_and_bulk_save(self):
        """Test custom column parent/child relationships and bulk save changes endpoint"""
        # Login admin first
        self.app.post('/register', data=dict(username='admin_test2', password='admin_password'))
        # Promote user to admin
        conn = database.get_db_connection()
        cursor = conn.cursor()
        user_row = cursor.execute("SELECT LoginId FROM Refusers WHERE Username = 'admin_test2'").fetchone()
        self.assertIsNotNone(user_row)
        user_id = user_row['loginid'] if 'loginid' in user_row else user_row[0]
        cursor.execute("UPDATE UserRole SET RoleId = 1 WHERE LoginId = ?", (user_id,))
        conn.commit()
        conn.close()
        
        self.app.post('/login', data=dict(username='admin_test2', password='admin_password'))
        
        # 1. Create a custom column with a parent
        resp = self.app.post('/api/admin/excel-columns/create', data=json.dumps({
            'column_key': 'test_child_col',
            'column_label': 'Test Child Column',
            'target_type': 'expense',
            'is_enabled_import': 1,
            'is_enabled_export': 1,
            'is_required': 0,
            'display_order': 12,
            'parent_column_key': 'payment_method',
            'parent_trigger_value': 'Credit'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Verify it has parent config
        resp = self.app.get('/api/admin/excel-columns?target_type=expense')
        cols = json.loads(resp.data)
        child_col = next((c for c in cols if c['column_key'] == 'test_child_col'), None)
        self.assertIsNotNone(child_col)
        self.assertEqual(child_col['parent_column_key'], 'payment_method')
        self.assertEqual(child_col['parent_trigger_value'], 'Credit')
        
        # 2. Test save-all endpoint
        save_payload = {
            'type_key': 'import',
            'columns': [
                {
                    'column_key': 'test_child_col',
                    'target_type': 'expense',
                    'display_order': 45,
                    'is_enabled': 0
                }
            ]
        }
        resp = self.app.post('/api/admin/excel-columns/save-all', data=json.dumps(save_payload), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Verify changes were saved (is_enabled_import should be 0, is_enabled_export should still be 1, display_order should be 45)
        resp = self.app.get('/api/admin/excel-columns?target_type=expense')
        cols = json.loads(resp.data)
        child_col_updated = next((c for c in cols if c['column_key'] == 'test_child_col'), None)
        self.assertIsNotNone(child_col_updated)
        self.assertEqual(child_col_updated['display_order'], 45)
        self.assertEqual(child_col_updated['is_enabled_import'], 0)
        self.assertEqual(child_col_updated['is_enabled_export'], 1)
        
        # Clean up
        resp = self.app.post('/api/admin/excel-columns/delete', data=json.dumps({
            'column_key': 'test_child_col',
            'target_type': 'expense'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        self.app.get('/logout')

    def test_mfa_and_otp_workflow(self):
        """Test MFA login, profile fields registration, and OTP send/verify endpoints"""
        # 1. Test OTP send and verification
        email = 'new_otp_user@example.com'
        phone = '9876543210'
        
        # Send email OTP
        resp = self.app.post('/api/otp/send', data=json.dumps({'target': email}), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Send phone OTP
        resp = self.app.post('/api/otp/send', data=json.dumps({'target': phone}), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Fetch OTPs from database to verify
        conn = database.get_db_connection()
        email_otp = conn.cursor().execute("SELECT otp FROM otps WHERE target = ?", (email,)).fetchone()[0]
        phone_otp = conn.cursor().execute("SELECT otp FROM otps WHERE target = ?", (phone,)).fetchone()[0]
        conn.close()
        
        # Verify incorrect email OTP
        resp = self.app.post('/api/otp/verify', data=json.dumps({'target': email, 'otp_code': '000000'}), content_type='application/json')
        self.assertEqual(resp.status_code, 400)
        
        # Verify correct email OTP
        resp = self.app.post('/api/otp/verify', data=json.dumps({'target': email, 'otp_code': email_otp}), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Verify correct phone OTP
        resp = self.app.post('/api/otp/verify', data=json.dumps({'target': phone, 'otp_code': phone_otp}), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # 2. Register user with profile fields
        resp = self.app.post('/register', data=json.dumps({
            'username': 'otp_test_user',
            'password': 'password123',
            'first_name': 'Test',
            'last_name': 'OTP',
            'email': email,
            'phone': phone
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        
        # Verify user is created in database with profile fields
        conn = database.get_db_connection()
        db_user = conn.cursor().execute("SELECT Username as username, Firstname as first_name, Lastname as last_name, Email as email, Phone as phone FROM Refusers WHERE Username = ?", ('otp_test_user',)).fetchone()
        conn.close()
        self.assertEqual(db_user['first_name'], 'Test')
        self.assertEqual(db_user['last_name'], 'OTP')
        self.assertEqual(db_user['email'], email)
        self.assertEqual(db_user['phone'], phone)
        
        # Log out the user from the auto-login session of registration
        self.app.get('/logout')
        
        # 3. Test MFA Login Flow
        # Disable app.testing temporarily to trigger real MFA flow
        app.testing = False
        old_db_path = os.environ.get('DATABASE_PATH')
        if 'DATABASE_PATH' in os.environ:
            del os.environ['DATABASE_PATH']
        try:
            database.set_setting('login_otp_enabled', '1')
            resp = self.app.post('/login', data=json.dumps({
                'username': 'otp_test_user',
                'password': 'password123'
            }), content_type='application/json')
            self.assertEqual(resp.status_code, 200)
            
            login_data = json.loads(resp.data)
            self.assertTrue(login_data.get('mfa_required'))
            temp_token = login_data.get('temp_token')
            self.assertIsNotNone(temp_token)
            
            # Fetch login OTP from database
            conn = database.get_db_connection()
            login_otp = conn.cursor().execute("SELECT otp FROM otps WHERE target = ?", (phone,)).fetchone()[0]
            conn.close()
            
            # Verify MFA login with incorrect OTP
            resp = self.app.post('/api/login/mfa', data=json.dumps({
                'temp_token': temp_token,
                'otp_code': '000000'
            }), content_type='application/json')
            self.assertEqual(resp.status_code, 400)
            
            # Verify MFA login with correct OTP
            resp = self.app.post('/api/login/mfa', data=json.dumps({
                'temp_token': temp_token,
                'otp_code': login_otp
            }), content_type='application/json')
            self.assertEqual(resp.status_code, 200)
        finally:
            app.testing = True
            if old_db_path is not None:
                os.environ['DATABASE_PATH'] = old_db_path
            
        self.app.get('/logout')

    def test_bulk_delete_endpoints(self):
        # Log in standard user
        self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        
        # 1. Create expenses
        e1_id = database.add_expense(1, 100.0, 'Food & Dining', 'e1', '2026-07-06', payment_method='Debit', status='Paid')
        e2_id = database.add_expense(1, 200.0, 'Shopping', 'e2', '2026-07-06', payment_method='Debit', status='Paid')
        
        # Verify they exist
        self.assertIsNotNone(e1_id)
        self.assertIsNotNone(e2_id)
        
        # Bulk delete them
        resp = self.app.post('/api/expenses/delete-bulk', data=json.dumps({
            'expense_ids': [e1_id, e2_id]
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        res_data = json.loads(resp.data)
        self.assertTrue(res_data.get('success'))
        
        # Verify they are deleted
        conn = database.get_db_connection()
        c = conn.cursor()
        e1 = c.execute('SELECT * FROM expenses WHERE id = ?', (e1_id,)).fetchone()
        e2 = c.execute('SELECT * FROM expenses WHERE id = ?', (e2_id,)).fetchone()
        conn.close()
        self.assertIsNone(e1)
        self.assertIsNone(e2)
        
        # 2. Create EMIs
        emi1_id = database.add_emi(1, 'emi1', 1000.0, 100.0, '2026-07-06', '2027-07-06', 12, 5.0, '5', 'Auto', 'GPay', 'SBI')
        emi2_id = database.add_emi(1, 'emi2', 2000.0, 200.0, '2026-07-06', '2027-07-06', 12, 5.0, '5', 'Auto', 'PhonePe', 'Kotak')
        
        # Verify they exist
        self.assertIsNotNone(emi1_id)
        self.assertIsNotNone(emi2_id)
        
        # Bulk delete them
        resp = self.app.post('/api/emis/delete-bulk', data=json.dumps({
            'emi_ids': [emi1_id, emi2_id]
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        res_data = json.loads(resp.data)
        self.assertTrue(res_data.get('success'))
        
        # Verify they are deleted
        conn = database.get_db_connection()
        c = conn.cursor()
        emi1 = c.execute('SELECT * FROM emis WHERE id = ?', (emi1_id,)).fetchone()
        emi2 = c.execute('SELECT * FROM emis WHERE id = ?', (emi2_id,)).fetchone()
        conn.close()
        self.assertIsNone(emi1)
        self.assertIsNone(emi2)
        
        self.app.get('/logout')

    def test_user_specific_features(self):
        # 1. Register a new user
        # Log in as admin to update settings
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Set separate registration OTP flags
        resp = self.app.post('/api/admin/settings/update', data=json.dumps({
            'register_email_otp_enabled': True,
            'register_phone_otp_enabled': False,
            'login_otp_enabled': False
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        self.app.get('/logout')

        # Try to register without email (should fail because email OTP is enabled)
        resp = self.app.post('/register', data=json.dumps({
            'username': 'otpuser1',
            'password': 'password123',
            'first_name': 'OTP',
            'last_name': 'User',
            'test_validation': True
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 400)

        # Register successfully with email
        resp = self.app.post('/register', data=json.dumps({
            'username': 'otpuser1',
            'password': 'password123',
            'first_name': 'OTP',
            'last_name': 'User',
            'email': 'otpuser1@example.com',
            'test_validation': True
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Get categories (should get child table categories)
        resp = self.app.get('/api/categories')
        self.assertEqual(resp.status_code, 200)
        cats = json.loads(resp.data)
        self.assertTrue(len(cats) > 0)

        # Create a new custom category for this user
        resp = self.app.post('/api/admin/categories/create', data=json.dumps({
            'name': 'CustomCategory1',
            'display_order': 100
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # Verify it exists in user's categories
        resp = self.app.get('/api/categories')
        cats = json.loads(resp.data)
        cat_names = [c['name'] for c in cats]
        self.assertIn('CustomCategory1', cat_names)

        # Log out, log in as admin, check admin categories (should NOT contain CustomCategory1)
        self.app.get('/logout')
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        resp = self.app.get('/api/categories')
        admin_cats = json.loads(resp.data)
        admin_cat_names = [c['name'] for c in admin_cats]
        self.assertNotIn('CustomCategory1', admin_cat_names)

        # Verify active currency is user-specific
        # Admin sets active currency to USA ($ - id 2)
        resp = self.app.post('/api/admin/currencies/set_active/2')
        self.assertEqual(resp.status_code, 200)

        # Admin active currency is now $
        resp = self.app.get('/api/active-currency')
        self.assertEqual(json.loads(resp.data)['symbol'], '$')

        # Log out, log in as otpuser1 (their active currency should fallback to India Rupee as it was never set)
        self.app.get('/logout')
        resp = self.app.post('/login', data=json.dumps({
            'username': 'otpuser1',
            'password': 'password123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        # User active currency is ₹ (fallback)
        resp = self.app.get('/api/active-currency')
        self.assertEqual(json.loads(resp.data)['symbol'], '₹')

        # User sets active currency to Europe (€ - id 3)
        resp = self.app.post('/api/admin/currencies/set_active/3')
        self.assertEqual(resp.status_code, 200)

        # User active currency is now €
        resp = self.app.get('/api/active-currency')
        self.assertEqual(json.loads(resp.data)['symbol'], '€')

        # Log out, log in as admin, admin's active currency is still $
        self.app.get('/logout')
        resp = self.app.post('/login', data=json.dumps({
            'username': 'admin',
            'password': 'admin123'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

        resp = self.app.get('/api/active-currency')
        self.assertEqual(json.loads(resp.data)['symbol'], '$')

        # Verify bulk additions endpoint
        resp = self.app.post('/api/expenses/add', data=json.dumps([
            {
                'amount': 100.0,
                'category': 'Food',
                'date': '2026-07-10',
                'description': 'Bulk item 1'
            },
            {
                'amount': 250.5,
                'category': 'Rent',
                'date': '2026-07-10',
                'description': 'Bulk item 2'
            }
        ]), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        bulk_res = json.loads(resp.data)
        self.assertTrue(bulk_res['success'])
        self.assertEqual(len(bulk_res['ids']), 2)

        # Test Chatbot API endpoint
        # First, query chatbot when logged in as admin
        resp = self.app.post('/api/chat', data=json.dumps({
            'message': 'total spent this month'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('reply', json.loads(resp.data))

        # Query chatbot for category breakdown
        resp = self.app.post('/api/chat', data=json.dumps({
            'message': 'category breakdown this month'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('reply', json.loads(resp.data))

        # Query chatbot for credit spending
        resp = self.app.post('/api/chat', data=json.dumps({
            'message': 'credit spending this month'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('reply', json.loads(resp.data))

        self.app.get('/logout')

        # Test chatbot API unauthorized when logged out
        resp = self.app.post('/api/chat', data=json.dumps({
            'message': 'total spent'
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 401)

if __name__ == '__main__':
    unittest.main()


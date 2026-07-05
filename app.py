import os
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import database

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'spendsmart-default-secret-key-12345')

# Initialize database
database.init_db()

# Helper to check login status
def is_logged_in():
    return 'user_id' in session

from functools import wraps

def require_privilege(privilege_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not is_logged_in():
                return jsonify({'error': 'Unauthorized'}), 401
            privs = database.get_user_privileges(session['user_id'])
            if privilege_name == 'can_admin':
                if not privs.get('is_admin'):
                    return jsonify({'error': 'Forbidden: Administrator privileges required'}), 403
            else:
                if not privs.get(privilege_name):
                    return jsonify({'error': f'Forbidden: Missing privilege {privilege_name}'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/')
def index():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template('index.html', username=session.get('username'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if is_logged_in():
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        # API-based login (handles JSON or form data)
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password are required.'}), 400
            
        user = database.get_user_by_username(username)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            return jsonify({'success': True, 'message': 'Logged in successfully.'})
        else:
            return jsonify({'success': False, 'message': 'Invalid username or password.'}), 401
            
    return render_template('login.html')

@app.route('/register', methods=['POST'])
def register():
    if is_logged_in():
        return jsonify({'success': False, 'message': 'Already logged in.'}), 400
        
    data = request.get_json() if request.is_json else request.form
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password are required.'}), 400
        
    if len(username) < 3:
        return jsonify({'success': False, 'message': 'Username must be at least 3 characters.'}), 400
        
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'}), 400
        
    hashed_password = generate_password_hash(password)
    user_id = database.create_user(username, hashed_password)
    
    if user_id:
        # Auto-login after registration
        session['user_id'] = user_id
        session['username'] = username
        return jsonify({'success': True, 'message': 'Registered and logged in successfully.'})
    else:
        return jsonify({'success': False, 'message': 'Username already exists.'}), 409

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/verify-username', methods=['GET', 'POST'])
def verify_username():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
    else:
        username = request.args.get('username', '').strip()
        
    if not username:
        return jsonify({'exists': False, 'message': 'Username is required.'}), 400
        
    user = database.get_user_by_username(username)
    if user:
        return jsonify({'exists': True})
    else:
        return jsonify({'exists': False})

@app.route('/api/change-password', methods=['POST'])
def change_password():
    data = request.get_json() if request.is_json else request.form
    username = data.get('username', '').strip()
    new_password = data.get('new_password', '')
    confirm_password = data.get('confirm_password', '')
    
    if not username or not new_password or not confirm_password:
        return jsonify({'success': False, 'message': 'Username, new password, and confirmation are required.'}), 400
        
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'New password must be at least 6 characters.'}), 400
        
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'Passwords do not match.'}), 400
        
    user = database.get_user_by_username(username)
    if not user:
        return jsonify({'success': False, 'message': 'Username does not exist.'}), 404
        
    hashed_password = generate_password_hash(new_password)
    success = database.update_user_password_by_username(username, hashed_password)
    
    if success:
        return jsonify({'success': True, 'message': 'Password changed successfully.'})
    else:
        return jsonify({'success': False, 'message': 'Failed to update password.'}), 500

# USER PRIVILEGES CHECK & CATEGORIES API (OPEN TO LOGGED IN USERS)

@app.route('/api/user/privileges', methods=['GET'])
def get_user_privileges_api():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    privs = database.get_user_privileges(session['user_id'])
    return jsonify({
        'username': session['username'],
        'privileges': privs
    })

@app.route('/api/categories', methods=['GET'])
def get_categories_api():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    cats = database.get_categories()
    return jsonify(cats)

# EXPENSE API ENDPOINTS WITH RBAC

@app.route('/api/expenses', methods=['GET'])
@require_privilege('can_view')
def get_expenses_api():
    category = request.args.get('category')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    bank_mode = request.args.get('bank_mode')
    payment_type = request.args.get('payment_type')
    payment_category = request.args.get('payment_category')
    payment_method = request.args.get('payment_method')
    status = request.args.get('status')
    month = request.args.get('month')
    year = request.args.get('year')
    
    expenses = database.get_expenses(
        session['user_id'],
        category=category,
        start_date=start_date,
        end_date=end_date,
        search=search,
        bank_mode=bank_mode,
        payment_type=payment_type,
        payment_category=payment_category,
        month=month,
        year=year,
        payment_method=payment_method,
        status=status
    )
    return jsonify(expenses)

@app.route('/api/expenses/add', methods=['POST'])
@require_privilege('can_add')
def add_expense_api():
    data = request.get_json() or {}
    amount = data.get('amount')
    category = data.get('category')
    description = data.get('description', '')
    date = data.get('date')
    bank_mode = data.get('bank_mode', '')
    payment_type = data.get('payment_type', '')
    payment_category = data.get('payment_category', '')
    payment_method = data.get('payment_method', 'Debit')
    interest = data.get('interest', 0.0)
    status = data.get('status', 'Paid')
    
    # Validations
    if amount is None or not category or not date:
        return jsonify({'error': 'Amount, Category, and Date are required.'}), 400
        
    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than zero.'}), 400
    except ValueError:
        return jsonify({'error': 'Amount must be a number.'}), 400

    try:
        interest = float(interest)
        if interest < 0:
            return jsonify({'error': 'Interest cannot be negative.'}), 400
    except ValueError:
        interest = 0.0
        
    # Extract dynamic custom fields
    standard_keys = {'amount', 'category', 'date', 'description', 'bank_mode', 'payment_type', 'payment_category', 'interest', 'payment_method', 'status'}
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    expense_id = database.add_expense(
        session['user_id'], amount, category, description, date,
        bank_mode=bank_mode, payment_type=payment_type, payment_category=payment_category,
        interest=interest, payment_method=payment_method, status=status,
        **extra_fields
    )
    return jsonify({'success': True, 'id': expense_id, 'message': 'Expense added successfully.'})

@app.route('/api/expenses/edit/<int:expense_id>', methods=['POST'])
@require_privilege('can_edit')
def edit_expense_api(expense_id):
    data = request.get_json() or {}
    amount = data.get('amount')
    category = data.get('category')
    description = data.get('description', '')
    date = data.get('date')
    bank_mode = data.get('bank_mode', '')
    payment_type = data.get('payment_type', '')
    payment_category = data.get('payment_category', '')
    payment_method = data.get('payment_method', 'Debit')
    interest = data.get('interest', 0.0)
    status = data.get('status', 'Paid')
    
    # Validations
    if amount is None or not category or not date:
        return jsonify({'error': 'Amount, Category, and Date are required.'}), 400
        
    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than zero.'}), 400
    except ValueError:
        return jsonify({'error': 'Amount must be a number.'}), 400

    try:
        interest = float(interest)
        if interest < 0:
            return jsonify({'error': 'Interest cannot be negative.'}), 400
    except ValueError:
        interest = 0.0
        
    # Extract dynamic custom fields
    standard_keys = {'amount', 'category', 'date', 'description', 'bank_mode', 'payment_type', 'payment_category', 'interest', 'payment_method', 'status'}
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    success = database.update_expense(
        expense_id, session['user_id'], amount, category, description, date,
        bank_mode=bank_mode, payment_type=payment_type, payment_category=payment_category,
        interest=interest, payment_method=payment_method, status=status,
        **extra_fields
    )
    if success:
        return jsonify({'success': True, 'message': 'Expense updated successfully.'})
    else:
        return jsonify({'error': 'Expense not found or update failed.'}), 404

@app.route('/api/expenses/delete/<int:expense_id>', methods=['POST', 'DELETE'])
@require_privilege('can_delete')
def delete_expense_api(expense_id):
    success = database.delete_expense(expense_id, session['user_id'])
    if success:
        return jsonify({'success': True, 'message': 'Expense deleted successfully.'})
    else:
        return jsonify({'error': 'Expense not found or deletion failed.'}), 404

@app.route('/api/overview', methods=['GET'])
@require_privilege('can_view')
def get_overview_api():
    month = request.args.get('month')
    year = request.args.get('year')
    data = database.get_overview_data(session['user_id'], month=month, year=year)
    return jsonify(data)

# EMI API ENDPOINTS

@app.route('/api/emis', methods=['GET'])
@require_privilege('can_view')
def get_emis_api():
    emis = database.get_emis(session['user_id'])
    return jsonify(emis)

@app.route('/api/emis/export', methods=['GET'])
@require_privilege('can_view')
def export_emis():
    user_id = session['user_id']
    emis = database.get_emis(user_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "EMIs"
    
    # Get active columns for EMI
    db_cols = database.get_excel_columns('emi')
    active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
    
    # Headers
    headers = [c['column_label'] for c in active_cols]
    ws.append(headers)
    
    for emi in emis:
        row_data = []
        for col in active_cols:
            k = col['column_key']
            if k == 'name':
                row_data.append(emi.get('name', ''))
            elif k == 'principal_amount':
                row_data.append(float(emi.get('principal_amount', 0.0)))
            elif k == 'interest_rate':
                row_data.append(float(emi.get('interest_rate', 0.0)))
            elif k == 'tenure_months':
                row_data.append(int(emi.get('tenure_months', 0)))
            elif k == 'emi_amount':
                row_data.append(float(emi.get('emi_amount', 0.0)))
            elif k == 'start_date':
                row_data.append(emi.get('start_date', ''))
            elif k == 'end_date':
                row_data.append(emi.get('end_date', ''))
            elif k == 'due_date':
                row_data.append(emi.get('due_date', ''))
            elif k == 'payment_type':
                row_data.append(emi.get('payment_type', ''))
            elif k == 'payment_gateway':
                row_data.append(emi.get('payment_gateway', ''))
            elif k == 'payment_bank':
                row_data.append(emi.get('payment_bank', ''))
            else:
                # Dynamic custom column
                row_data.append(emi.get(k, ''))
        ws.append(row_data)
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    download_name = f"emis_export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name
    )

@app.route('/api/emis/import', methods=['POST'])
@require_privilege('can_add')
def import_emis():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
        
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            return jsonify({'error': 'The uploaded Excel file is empty'}), 400
            
        header_map = {}
        for i, val in enumerate(first_row):
            if val is not None:
                header_map[str(val).strip().lower()] = i
                
        # Get active columns configuration (specifically for Import)
        db_cols = database.get_excel_columns('emi')
        enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
        
        # Verify required headers are present
        missing_reqs = []
        for k, req in enabled_keys.items():
            if req:
                label = next(c['column_label'] for c in db_cols if c['column_key'] == k)
                if label.lower() not in header_map:
                    missing_reqs.append(label)
        if missing_reqs:
            return jsonify({'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'}), 400
            
        imported_count = 0
        user_id = session['user_id']
        
        for r_idx, row in enumerate(rows_iter, start=2):
            if not any(val is not None for val in row):
                continue
                
            # Map fields dynamically from columns list
            data_dict = {}
            for col in db_cols:
                k = col['column_key']
                if col['is_enabled_import'] != 1:
                    continue
                label = col['column_label'].lower()
                idx = header_map.get(label) or header_map.get(k)
                if idx is not None:
                    data_dict[k] = row[idx]
            
            name = str(data_dict.get('name') or '').strip()
            principal = float(data_dict.get('principal_amount') or 0.0)
            interest = float(data_dict.get('interest_rate') or 0.0)
            tenure = int(data_dict.get('tenure_months') or 12)
            emi_amount = float(data_dict.get('emi_amount') or 0.0)
            
            start_date_val = data_dict.get('start_date')
            end_date_val = data_dict.get('end_date')
            
            if isinstance(start_date_val, (datetime.datetime, datetime.date)):
                start_date = start_date_val.strftime('%Y-%m-%d')
            else:
                start_date = str(start_date_val or '').strip()
                
            if isinstance(end_date_val, (datetime.datetime, datetime.date)):
                end_date = end_date_val.strftime('%Y-%m-%d')
            else:
                end_date = str(end_date_val or '').strip()
                
            due_date = str(data_dict.get('due_date') or '5').strip()
            payment_type = str(data_dict.get('payment_type') or 'Auto').strip()
            gateway = str(data_dict.get('payment_gateway') or '').strip()
            bank = str(data_dict.get('payment_bank') or '').strip()
            
            if not name or not start_date or not end_date or not due_date or not payment_type:
                continue
                
            # Extra fields extraction
            extra_fields = {}
            for k in data_dict:
                if k not in ('name', 'principal_amount', 'interest_rate', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'):
                    extra_fields[k] = str(data_dict[k]).strip() if data_dict[k] is not None else ''
                    
            database.add_emi(
                user_id, name, principal, emi_amount, start_date, end_date, tenure, interest, due_date, payment_type, gateway, bank,
                **extra_fields
            )
            imported_count += 1
            
        return jsonify({'success': True, 'message': f'Successfully imported {imported_count} EMIs.'})
        
    except Exception as e:
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

@app.route('/api/emis/add', methods=['POST'])
@require_privilege('can_add')
def add_emi_api():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    principal_amount = data.get('principal_amount', 0.0)
    emi_amount = data.get('emi_amount')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    tenure_months = data.get('tenure_months')
    interest_rate = data.get('interest_rate', 0.0)
    due_date = data.get('due_date', '').strip()
    payment_type = data.get('payment_type', '').strip()
    payment_gateway = data.get('payment_gateway', '').strip()
    payment_bank = data.get('payment_bank', '').strip()
    
    if not name or emi_amount is None or not start_date or not end_date or not tenure_months or not due_date or not payment_type:
        return jsonify({'error': 'Name, EMI Amount, Start/End Date, Tenure, Due Date, and Payment Type are required.'}), 400
        
    try:
        principal_amount = float(principal_amount)
        emi_amount = float(emi_amount)
        interest_rate = float(interest_rate)
        tenure_months = int(tenure_months)
    except ValueError:
        return jsonify({'error': 'Numeric fields must contain valid numbers.'}), 400
        
    standard_keys = {
        'name', 'principal_amount', 'emi_amount', 'start_date', 'end_date', 
        'tenure_months', 'interest_rate', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'
    }
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    emi_id = database.add_emi(
        session['user_id'], name, principal_amount, emi_amount, start_date, end_date, tenure_months, interest_rate, due_date, payment_type, payment_gateway, payment_bank,
        **extra_fields
    )
    return jsonify({'success': True, 'id': emi_id, 'message': 'EMI added successfully.'})

@app.route('/api/emis/edit/<int:emi_id>', methods=['POST'])
@require_privilege('can_edit')
def edit_emi_api(emi_id):
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    principal_amount = data.get('principal_amount', 0.0)
    emi_amount = data.get('emi_amount')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    tenure_months = data.get('tenure_months')
    interest_rate = data.get('interest_rate', 0.0)
    due_date = data.get('due_date', '').strip()
    payment_type = data.get('payment_type', '').strip()
    payment_gateway = data.get('payment_gateway', '').strip()
    payment_bank = data.get('payment_bank', '').strip()
    
    if not name or emi_amount is None or not start_date or not end_date or not tenure_months or not due_date or not payment_type:
        return jsonify({'error': 'Name, EMI Amount, Start/End Date, Tenure, Due Date, and Payment Type are required.'}), 400
        
    try:
        principal_amount = float(principal_amount)
        emi_amount = float(emi_amount)
        interest_rate = float(interest_rate)
        tenure_months = int(tenure_months)
    except ValueError:
        return jsonify({'error': 'Numeric fields must contain valid numbers.'}), 400
        
    standard_keys = {
        'name', 'principal_amount', 'emi_amount', 'start_date', 'end_date', 
        'tenure_months', 'interest_rate', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'
    }
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    success = database.update_emi(
        emi_id, name, principal_amount, emi_amount, start_date, end_date, tenure_months, interest_rate, due_date, payment_type, payment_gateway, payment_bank, user_id=session['user_id'],
        **extra_fields
    )
    if success:
        return jsonify({'success': True, 'message': 'EMI updated successfully.'})
    else:
        return jsonify({'error': 'EMI not found or update failed.'}), 404

@app.route('/api/emis/delete/<int:emi_id>', methods=['POST', 'DELETE'])
@require_privilege('can_delete')
def delete_emi_api(emi_id):
    success = database.delete_emi(emi_id, session['user_id'])
    if success:
        return jsonify({'success': True, 'message': 'EMI deleted successfully.'})
    else:
        return jsonify({'error': 'EMI not found or deletion failed.'}), 404

# EXCEL IMPORT/EXPORT API ENDPOINTS
from io import BytesIO
from openpyxl import Workbook, load_workbook
import datetime

@app.route('/api/expenses/import-template', methods=['GET'])
def get_import_template():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    # Get active columns
    db_cols = database.get_excel_columns('expense')
    active_cols = [c for c in db_cols if c['is_enabled_import'] == 1]
    
    # Headers
    headers = [c['column_label'] for c in active_cols]
    ws.append(headers)
    
    # Sample placeholders
    placeholders = {
        'date': "2026-06-25",
        'category': "Food & Dining",
        'description': "Lunch with friends",
        'gateway': "GPay",
        'bank': "SBI",
        'source': "Salary",
        'method': "Debit",
        'amount': 12.50,
        'interest': 0.00,
        'status': "Paid"
    }
    placeholders2 = {
        'date': "2026-06-26",
        'category': "Shopping",
        'description': "Bought a laptop",
        'gateway': "Credit Card",
        'bank': "Kotak",
        'source': "Credit Card",
        'method': "Credit",
        'amount': 1200.00,
        'interest': 45.00,
        'status': "Unpaid"
    }
    
    # Add sample placeholder rows
    row1 = [placeholders.get(c['column_key'], '') for c in active_cols]
    row2 = [placeholders2.get(c['column_key'], '') for c in active_cols]
    ws.append(row1)
    ws.append(row2)
    
    # Adjust column widths for readability
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="spendsmart_import_template.xlsx"
    )

@app.route('/api/expenses/export', methods=['GET'])
@require_privilege('can_view')
def export_expenses():
    month = request.args.get('month')
    year = request.args.get('year')
    
    # Only filter by month and year, ignore other filters per instructions
    expenses = database.get_expenses(
        session['user_id'],
        month=month,
        year=year
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Get active columns
    db_cols = database.get_excel_columns('expense')
    active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
    
    # Headers
    headers = [c['column_label'] for c in active_cols]
    ws.append(headers)
    
    # Write data
    for exp in expenses:
        row_data = []
        for col in active_cols:
            k = col['column_key']
            if k == 'date':
                row_data.append(exp.get('date', ''))
            elif k == 'category':
                row_data.append(exp.get('category', ''))
            elif k == 'description':
                row_data.append(exp.get('description', ''))
            elif k == 'gateway':
                row_data.append(exp.get('payment_type', ''))
            elif k == 'bank':
                row_data.append(exp.get('bank_mode', ''))
            elif k == 'source':
                row_data.append(exp.get('payment_category', ''))
            elif k == 'method':
                row_data.append(exp.get('payment_method', 'Debit'))
            elif k == 'amount':
                row_data.append(float(exp.get('amount', 0.0)))
            elif k == 'interest':
                row_data.append(float(exp.get('interest', 0.0)))
            elif k == 'status':
                row_data.append(exp.get('status', 'Paid'))
            else:
                # Dynamic custom column
                row_data.append(exp.get(k, ''))
        ws.append(row_data)
        
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    download_name = f"expenses_export"
    if year:
        download_name += f"_{year}"
    if month:
        download_name += f"_{month}"
    download_name += f"_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name
    )

@app.route('/api/expenses/import', methods=['POST'])
@require_privilege('can_add')
def import_expenses():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
        
    try:
        # Load workbook (in-memory bytes)
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            return jsonify({'error': 'The uploaded Excel file is empty'}), 400
            
        # Locate columns in header
        header_map = {}
        for i, val in enumerate(first_row):
            if val is not None:
                header_map[str(val).strip().lower()] = i
                
        # Get active columns configuration (specifically for Import)
        db_cols = database.get_excel_columns('expense')
        enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
        
        # Required columns mapping
        col_date = header_map.get('date') if 'date' in enabled_keys else None
        col_category = header_map.get('category') if 'category' in enabled_keys else None
        col_amount = header_map.get('amount') if 'amount' in enabled_keys else None
        
        # Verify required headers are present (only if enabled)
        missing_reqs = []
        if 'date' in enabled_keys and col_date is None: missing_reqs.append('Date')
        if 'category' in enabled_keys and col_category is None: missing_reqs.append('Category')
        if 'amount' in enabled_keys and col_amount is None: missing_reqs.append('Amount')
        
        if missing_reqs:
            return jsonify({
                'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'
            }), 400
            
        # Optional columns mapping (if enabled)
        col_desc = header_map.get('description') if 'description' in enabled_keys else None
        col_gateway = (header_map.get('gateway') or header_map.get('payment type') or header_map.get('payment_type')) if 'gateway' in enabled_keys else None
        col_bank = (header_map.get('bank') or header_map.get('bank mode') or header_map.get('bank_mode')) if 'bank' in enabled_keys else None
        col_source = (header_map.get('source') or header_map.get('payment category') or header_map.get('payment_category')) if 'source' in enabled_keys else None
        col_method = (header_map.get('method') or header_map.get('payment method') or header_map.get('payment_method')) if 'method' in enabled_keys else None
        col_interest = header_map.get('interest') if 'interest' in enabled_keys else None
        col_status = header_map.get('status') if 'status' in enabled_keys else None
        
        # Let's map any other active column dynamically:
        dynamic_cols = {}
        for col in db_cols:
            k = col['column_key']
            if col['is_enabled_import'] == 1 and k not in ('date', 'category', 'description', 'gateway', 'bank', 'source', 'method', 'amount', 'interest', 'status'):
                # Locate it in the header either by key or label
                col_idx = header_map.get(k) or header_map.get(col['column_label'].lower())
                if col_idx is not None:
                    dynamic_cols[k] = col_idx
                    
        imported_count = 0
        skipped_count = 0
        
        for r_idx, row in enumerate(rows_iter, start=2):
            # Skip empty rows
            if not any(val is not None for val in row):
                continue
                
            date_val = row[col_date] if col_date is not None else None
            category_val = row[col_category] if col_category is not None else None
            amount_val = row[col_amount] if col_amount is not None else None
            
            # Date validation and normalization (fallback if disabled)
            if 'date' in enabled_keys and col_date is not None:
                if date_val is None:
                    skipped_count += 1
                    continue
                if isinstance(date_val, (datetime.datetime, datetime.date)):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val).strip()
                    try:
                        datetime.datetime.strptime(date_str, '%Y-%m-%d')
                    except ValueError:
                        skipped_count += 1
                        continue
            else:
                date_str = datetime.date.today().strftime('%Y-%m-%d')
                
            # Category validation and normalization (fallback if disabled)
            if 'category' in enabled_keys and col_category is not None:
                if category_val is None:
                    skipped_count += 1
                    continue
                category_val = str(category_val).strip()
                if not category_val:
                    skipped_count += 1
                    continue
            else:
                category_val = 'Other'
                
            # Amount normalization (fallback if disabled)
            if 'amount' in enabled_keys and col_amount is not None:
                if amount_val is None:
                    skipped_count += 1
                    continue
                try:
                    amount = float(amount_val)
                    if amount <= 0:
                        skipped_count += 1
                        continue
                except (ValueError, TypeError):
                    skipped_count += 1
                    continue
            else:
                amount = 0.01
                
            # Optional parameters
            description = str(row[col_desc]).strip() if (col_desc is not None and row[col_desc] is not None) else ''
            gateway = str(row[col_gateway]).strip() if (col_gateway is not None and row[col_gateway] is not None) else ''
            bank = str(row[col_bank]).strip() if (col_bank is not None and row[col_bank] is not None) else ''
            source = str(row[col_source]).strip() if (col_source is not None and row[col_source] is not None) else ''
            method = str(row[col_method]).strip() if (col_method is not None and row[col_method] is not None) else 'Debit'
            
            # Ensure method is normalized to 'Debit' or 'Credit'
            if method.lower() in ('credit', 'c'):
                method = 'Credit'
            else:
                method = 'Debit'
                
            # Interest validation
            interest = 0.0
            if col_interest is not None and row[col_interest] is not None:
                try:
                    interest = float(row[col_interest])
                    if interest < 0:
                        interest = 0.0
                except (ValueError, TypeError):
                    interest = 0.0
                    
            # If method is Debit, override interest to 0.0 per business rules
            if method == 'Debit':
                interest = 0.0
                
            status_val = 'Paid'
            if col_status is not None and row[col_status] is not None:
                status_raw = str(row[col_status]).strip().lower()
                if status_raw in ('unpaid', 'u'):
                    status_val = 'Unpaid'
                    
            # Extra dynamic fields extraction
            extra_fields = {}
            for k, col_idx in dynamic_cols.items():
                val = row[col_idx]
                extra_fields[k] = str(val).strip() if val is not None else ''

            database.add_expense(
                session['user_id'], amount, category_val, description, date_str,
                bank_mode=bank, payment_type=gateway, payment_category=source,
                interest=interest, payment_method=method, status=status_val,
                **extra_fields
            )
            imported_count += 1
            
        return jsonify({
            'success': True,
            'message': f'Import completed. {imported_count} expenses imported successfully. {skipped_count} invalid rows skipped.',
            'imported': imported_count,
            'skipped': skipped_count
        })
    except Exception as e:
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

# EXCEL COLUMNS CONFIGURATION CRUD ENDPOINTS (Open to all logged-in users for viewing, toggle/creation/deletion restricted to Admin)
@app.route('/api/admin/excel-columns', methods=['GET'])
def admin_get_excel_columns():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    target_type = request.args.get('target_type', 'expense')
    cols = database.get_excel_columns(target_type)
    return jsonify(cols)

@app.route('/api/admin/excel-columns/toggle', methods=['POST'])
@require_privilege('can_admin')
def admin_toggle_excel_column():
    data = request.get_json() or {}
    column_key = data.get('column_key')
    is_enabled = data.get('is_enabled')
    type_key = data.get('type_key')
    target_type = data.get('target_type', 'expense')
    
    if column_key is None or is_enabled is None or type_key not in ('import', 'export'):
        return jsonify({'error': 'Column key, is_enabled, and valid type_key ("import" or "export") are required.'}), 400
        
    db_cols = database.get_excel_columns(target_type)
    target_col = next((c for c in db_cols if c['column_key'] == column_key), None)
    if not target_col:
        return jsonify({'error': 'Column not found.'}), 404
        
    database.update_excel_column_status(column_key, type_key, int(is_enabled), target_type)
    return jsonify({'success': True, 'message': 'Column status updated successfully.'})

@app.route('/api/admin/excel-columns/save-all', methods=['POST'])
@require_privilege('can_admin')
def admin_save_all_excel_columns():
    data = request.get_json() or {}
    columns_data = data.get('columns', [])
    type_key = data.get('type_key', 'import')
    
    if type_key not in ('import', 'export'):
        return jsonify({'error': 'Invalid type_key ("import" or "export")'}), 400
        
    conn = database.get_db_connection()
    cursor = conn.cursor()
    field = 'is_enabled_import' if type_key == 'import' else 'is_enabled_export'
    
    for col in columns_data:
        col_key = col.get('column_key')
        target_type = col.get('target_type', 'expense')
        display_order = col.get('display_order', 0)
        is_enabled = col.get('is_enabled', 1)
        
        cursor.execute(
            f"UPDATE excel_columns SET display_order = ?, {field} = ? WHERE column_key = ? AND target_type = ?",
            (int(display_order), int(is_enabled), col_key, target_type)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'All configurations saved successfully.'})

@app.route('/api/admin/excel-columns/update-order', methods=['POST'])
@require_privilege('can_admin')
def admin_update_excel_column_order():
    data = request.get_json() or {}
    column_key = data.get('column_key')
    display_order = data.get('display_order')
    target_type = data.get('target_type', 'expense')
    
    if column_key is None or display_order is None:
        return jsonify({'error': 'Column key and display_order are required.'}), 400
        
    conn = database.get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE excel_columns SET display_order = ? WHERE column_key = ? AND target_type = ?",
        (int(display_order), column_key, target_type)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Column order updated successfully.'})

@app.route('/api/admin/excel-columns/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_excel_column():
    import sqlite3
    data = request.get_json() or {}
    column_key = data.get('column_key', '').strip().lower()
    column_label = data.get('column_label', '').strip()
    target_type = data.get('target_type', 'expense').strip().lower()
    is_required = int(data.get('is_required', 0))
    is_enabled_import = int(data.get('is_enabled_import', 1))
    is_enabled_export = int(data.get('is_enabled_export', 1))
    display_order = int(data.get('display_order', 0))
    parent_column_key = data.get('parent_column_key', '').strip() or None
    parent_trigger_value = data.get('parent_trigger_value', '').strip() or None
    
    if not column_key or not column_label or target_type not in ('expense', 'emi'):
        return jsonify({'error': 'Column key, label, and valid target_type ("expense" or "emi") are required.'}), 400
        
    import re
    if not re.match(r'^[a-z0-9_]+$', column_key):
        return jsonify({'error': 'Column key must only contain lowercase alphanumeric characters and underscores.'}), 400
        
    conn = database.get_db_connection()
    cursor = conn.cursor()
    try:
        # Dynamically alter table to add column if it doesn't exist
        table_name = 'expenses' if target_type == 'expense' else 'emis'
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [row['name'] for row in cursor.fetchall()]
        if column_key not in columns:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_key} TEXT")
            
        cursor.execute(
            'INSERT INTO excel_columns (column_key, column_label, is_enabled_import, is_enabled_export, is_required, target_type, display_order, parent_column_key, parent_trigger_value) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (column_key, column_label, is_enabled_import, is_enabled_export, is_required, target_type, display_order, parent_column_key, parent_trigger_value)
        )
        conn.commit()
        return jsonify({'success': True, 'message': f'Column registered and added to {table_name} table successfully.'})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Column key already exists for this target type.'}), 409
    except Exception as e:
        return jsonify({'error': f'Failed to register column: {str(e)}'}), 500
    finally:
        conn.close()

@app.route('/api/admin/excel-columns/delete', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_excel_column():
    data = request.get_json() or {}
    column_key = data.get('column_key')
    target_type = data.get('target_type', 'expense')
    
    if not column_key:
        return jsonify({'error': 'Column key is required.'}), 400
        
    system_keys = {
        'expense': ('date', 'category', 'amount'),
        'emi': ('name', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type')
    }
    if column_key in system_keys.get(target_type, ()):
        return jsonify({'error': 'Cannot delete system required columns.'}), 400
        
    conn = database.get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        'DELETE FROM excel_columns WHERE column_key = ? AND target_type = ?',
        (column_key, target_type)
    )
    conn.commit()
    rows_affected = cursor.rowcount
    conn.close()
    if rows_affected > 0:
        return jsonify({'success': True, 'message': 'Column removed successfully.'})
    else:
        return jsonify({'error': 'Column not found.'}), 404

# ADMIN API ENDPOINTS

@app.route('/api/admin/users', methods=['GET'])
@require_privilege('can_admin')
def admin_get_users():
    users = database.get_all_users()
    return jsonify(users)

@app.route('/api/admin/users/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role_id = data.get('role_id')
    
    if not username or not password or not role_id:
        return jsonify({'error': 'Username, password, and role are required.'}), 400
        
    if len(username) < 3:
        return jsonify({'error': 'Username must be at least 3 characters.'}), 400
        
    if len(password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters.'}), 400
        
    hashed_password = generate_password_hash(password)
    user_id = database.create_user(username, hashed_password, int(role_id))
    
    if user_id:
        return jsonify({'success': True, 'message': 'User created successfully.'})
    else:
        return jsonify({'error': 'Username already exists.'}), 409

@app.route('/api/admin/users/edit_role', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_user_role():
    data = request.get_json()
    user_id = data.get('user_id')
    role_id = data.get('role_id')
    
    if not user_id or not role_id:
        return jsonify({'error': 'User ID and Role ID are required.'}), 400
        
    # Prevent changing last admin's role
    current_privs = database.get_user_privileges(int(user_id))
    if current_privs.get('is_admin') and int(role_id) != 1:
        # Check if last admin
        users = database.get_all_users()
        admins = [u for u in users if u['role_id'] == 1]
        if len(admins) <= 1:
            return jsonify({'error': 'Cannot change role of the last administrator.'}), 400

    database.update_user_role(int(user_id), int(role_id))
    return jsonify({'success': True, 'message': 'User role updated successfully.'})

@app.route('/api/admin/users/change_password', methods=['POST'])
@require_privilege('can_admin')
def admin_change_user_password():
    data = request.get_json()
    user_id = data.get('user_id')
    new_password = data.get('new_password')
    confirm_password = data.get('confirm_password')
    
    if not user_id or not new_password or not confirm_password:
        return jsonify({'error': 'User ID, new password, and confirmation are required.'}), 400
        
    if len(new_password) < 6:
        return jsonify({'error': 'New password must be at least 6 characters.'}), 400
        
    if new_password != confirm_password:
        return jsonify({'error': 'Passwords do not match.'}), 400
        
    user = database.get_user_by_id(int(user_id))
    if not user:
        return jsonify({'error': 'User not found.'}), 404
        
    hashed_password = generate_password_hash(new_password)
    success = database.update_user_password(int(user_id), hashed_password)
    if success:
        return jsonify({'success': True, 'message': 'User password updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update user password.'}), 500

@app.route('/api/admin/users/delete/<int:user_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_user(user_id):
    success = database.delete_user(user_id)
    if success:
        return jsonify({'success': True, 'message': 'User deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete user. Ensure it is not the last Administrator.'}), 400

# ADMIN EMI API ENDPOINTS

@app.route('/api/admin/emis', methods=['GET'])
@require_privilege('can_admin')
def admin_get_emis():
    emis = database.get_all_emis()
    return jsonify(emis)

@app.route('/api/admin/emis/export', methods=['GET'])
@require_privilege('can_admin')
def admin_export_emis():
    emis = database.get_all_emis()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All EMIs"
    
    # Get active columns for EMI
    db_cols = database.get_excel_columns('emi')
    active_cols = [c for c in db_cols if c['is_enabled_export'] == 1]
    
    # Headers
    headers = ["Username"] + [c['column_label'] for c in active_cols]
    ws.append(headers)
    
    for emi in emis:
        row_data = [emi.get('username', '')]
        for col in active_cols:
            k = col['column_key']
            if k == 'name':
                row_data.append(emi.get('name', ''))
            elif k == 'principal_amount':
                row_data.append(float(emi.get('principal_amount', 0.0)))
            elif k == 'interest_rate':
                row_data.append(float(emi.get('interest_rate', 0.0)))
            elif k == 'tenure_months':
                row_data.append(int(emi.get('tenure_months', 0)))
            elif k == 'emi_amount':
                row_data.append(float(emi.get('emi_amount', 0.0)))
            elif k == 'start_date':
                row_data.append(emi.get('start_date', ''))
            elif k == 'end_date':
                row_data.append(emi.get('end_date', ''))
            elif k == 'due_date':
                row_data.append(emi.get('due_date', ''))
            elif k == 'payment_type':
                row_data.append(emi.get('payment_type', ''))
            elif k == 'payment_gateway':
                row_data.append(emi.get('payment_gateway', ''))
            elif k == 'payment_bank':
                row_data.append(emi.get('payment_bank', ''))
            else:
                # Dynamic custom column
                row_data.append(emi.get(k, ''))
        ws.append(row_data)
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    download_name = f"admin_emis_export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name
    )

@app.route('/api/admin/emis/import', methods=['POST'])
@require_privilege('can_admin')
def admin_import_emis():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx).'}), 400
        
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            return jsonify({'error': 'The uploaded Excel file is empty'}), 400
            
        header_map = {}
        for i, val in enumerate(first_row):
            if val is not None:
                header_map[str(val).strip().lower()] = i
                
        # Get active columns configuration (specifically for Import)
        db_cols = database.get_excel_columns('emi')
        enabled_keys = {c['column_key']: c['is_required'] for c in db_cols if c['is_enabled_import'] == 1}
        
        # Verify required headers are present
        if 'username' not in header_map:
            return jsonify({'error': 'Required column "Username" missing in Excel.'}), 400
            
        missing_reqs = []
        for k, req in enabled_keys.items():
            if req:
                label = next(c['column_label'] for c in db_cols if c['column_key'] == k)
                if label.lower() not in header_map:
                    missing_reqs.append(label)
        if missing_reqs:
            return jsonify({'error': f'Required columns missing in Excel: {", ".join(missing_reqs)}'}), 400
            
        imported_count = 0
        skipped_count = 0
        
        for r_idx, row in enumerate(rows_iter, start=2):
            if not any(val is not None for val in row):
                continue
                
            username = str(row[header_map["username"]] or '').strip()
            if not username:
                skipped_count += 1
                continue
                
            user = database.get_user_by_username(username)
            if not user:
                skipped_count += 1
                continue
                
            user_id = user['id']
            
            # Map fields dynamically from columns list
            data_dict = {}
            for col in db_cols:
                k = col['column_key']
                if col['is_enabled_import'] != 1:
                    continue
                label = col['column_label'].lower()
                idx = header_map.get(label) or header_map.get(k)
                if idx is not None:
                    data_dict[k] = row[idx]
            
            name = str(data_dict.get('name') or '').strip()
            principal = float(data_dict.get('principal_amount') or 0.0)
            interest = float(data_dict.get('interest_rate') or 0.0)
            tenure = int(data_dict.get('tenure_months') or 12)
            emi_amount = float(data_dict.get('emi_amount') or 0.0)
            
            start_date_val = data_dict.get('start_date')
            end_date_val = data_dict.get('end_date')
            
            if isinstance(start_date_val, (datetime.datetime, datetime.date)):
                start_date = start_date_val.strftime('%Y-%m-%d')
            else:
                start_date = str(start_date_val or '').strip()
                
            if isinstance(end_date_val, (datetime.datetime, datetime.date)):
                end_date = end_date_val.strftime('%Y-%m-%d')
            else:
                end_date = str(end_date_val or '').strip()
                
            due_date = str(data_dict.get('due_date') or '5').strip()
            payment_type = str(data_dict.get('payment_type') or 'Auto').strip()
            gateway = str(data_dict.get('payment_gateway') or '').strip()
            bank = str(data_dict.get('payment_bank') or '').strip()
            
            if not name or not start_date or not end_date or not due_date or not payment_type:
                skipped_count += 1
                continue
                
            # Extra fields extraction
            extra_fields = {}
            for k in data_dict:
                if k not in ('name', 'principal_amount', 'interest_rate', 'tenure_months', 'emi_amount', 'start_date', 'end_date', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'):
                    extra_fields[k] = str(data_dict[k]).strip() if data_dict[k] is not None else ''
                    
            database.add_emi(
                user_id, name, principal, emi_amount, start_date, end_date, tenure, interest, due_date, payment_type, gateway, bank,
                **extra_fields
            )
            imported_count += 1
            
        msg = f'Successfully imported {imported_count} EMIs.'
        if skipped_count > 0:
            msg += f' Skipped {skipped_count} rows due to invalid/missing username or data.'
            
        return jsonify({'success': True, 'message': msg})
        
    except Exception as e:
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

@app.route('/api/admin/emis/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_emi():
    data = request.get_json() or {}
    user_id = data.get('user_id')
    name = data.get('name', '').strip()
    principal_amount = data.get('principal_amount', 0.0)
    emi_amount = data.get('emi_amount')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    tenure_months = data.get('tenure_months')
    interest_rate = data.get('interest_rate', 0.0)
    due_date = data.get('due_date', '').strip()
    payment_type = data.get('payment_type', '').strip()
    payment_gateway = data.get('payment_gateway', '').strip()
    payment_bank = data.get('payment_bank', '').strip()
    
    if not user_id or not name or emi_amount is None or not start_date or not end_date or not tenure_months or not due_date or not payment_type:
        return jsonify({'error': 'User, Name, EMI Amount, Start/End Date, Tenure, Due Date, and Payment Type are required.'}), 400
        
    try:
        user_id = int(user_id)
        principal_amount = float(principal_amount)
        emi_amount = float(emi_amount)
        interest_rate = float(interest_rate)
        tenure_months = int(tenure_months)
    except ValueError:
        return jsonify({'error': 'Numeric fields must contain valid numbers.'}), 400
        
    standard_keys = {
        'user_id', 'name', 'principal_amount', 'emi_amount', 'start_date', 'end_date', 
        'tenure_months', 'interest_rate', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'
    }
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    emi_id = database.add_emi(
        user_id, name, principal_amount, emi_amount, start_date, end_date, tenure_months, interest_rate, due_date, payment_type, payment_gateway, payment_bank,
        **extra_fields
    )
    return jsonify({'success': True, 'id': emi_id, 'message': 'EMI created successfully.'})

@app.route('/api/admin/emis/edit/<int:emi_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_emi(emi_id):
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    principal_amount = data.get('principal_amount', 0.0)
    emi_amount = data.get('emi_amount')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    tenure_months = data.get('tenure_months')
    interest_rate = data.get('interest_rate', 0.0)
    due_date = data.get('due_date', '').strip()
    payment_type = data.get('payment_type', '').strip()
    payment_gateway = data.get('payment_gateway', '').strip()
    payment_bank = data.get('payment_bank', '').strip()
    
    if not name or emi_amount is None or not start_date or not end_date or not tenure_months or not due_date or not payment_type:
        return jsonify({'error': 'Name, EMI Amount, Start/End Date, Tenure, Due Date, and Payment Type are required.'}), 400
        
    try:
        principal_amount = float(principal_amount)
        emi_amount = float(emi_amount)
        interest_rate = float(interest_rate)
        tenure_months = int(tenure_months)
    except ValueError:
        return jsonify({'error': 'Numeric fields must contain valid numbers.'}), 400
        
    standard_keys = {
        'name', 'principal_amount', 'emi_amount', 'start_date', 'end_date', 
        'tenure_months', 'interest_rate', 'due_date', 'payment_type', 'payment_gateway', 'payment_bank'
    }
    extra_fields = {k: v for k, v in data.items() if k not in standard_keys}
        
    success = database.update_emi(
        emi_id, name, principal_amount, emi_amount, start_date, end_date, tenure_months, interest_rate, due_date, payment_type, payment_gateway, payment_bank,
        **extra_fields
    )
    if success:
        return jsonify({'success': True, 'message': 'EMI updated successfully.'})
    else:
        return jsonify({'error': 'EMI not found or update failed.'}), 404

@app.route('/api/admin/emis/delete/<int:emi_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_emi(emi_id):
    success = database.delete_emi(emi_id)
    if success:
        return jsonify({'success': True, 'message': 'EMI deleted successfully.'})
    else:
        return jsonify({'error': 'EMI not found or deletion failed.'}), 404

@app.route('/api/admin/roles', methods=['GET'])
@require_privilege('can_admin')
def admin_get_roles():
    roles = database.get_roles()
    return jsonify(roles)

@app.route('/api/admin/roles/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_role():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Role name is required.'}), 400
        
    role_id = database.add_role(name)
    if role_id:
        return jsonify({'success': True, 'id': role_id, 'message': 'Role created successfully.'})
    else:
        return jsonify({'error': 'Role already exists.'}), 409

@app.route('/api/admin/roles/delete/<int:role_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_role(role_id):
    success = database.delete_role(role_id)
    if success:
        return jsonify({'success': True, 'message': 'Role deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete role. Admin role cannot be deleted.'}), 400

@app.route('/api/admin/privileges/update', methods=['POST'])
@require_privilege('can_admin')
def admin_update_privileges():
    data = request.get_json()
    role_id = data.get('role_id')
    can_view = data.get('can_view', 0)
    can_add = data.get('can_add', 0)
    can_edit = data.get('can_edit', 0)
    can_delete = data.get('can_delete', 0)
    
    if not role_id:
        return jsonify({'error': 'Role ID is required.'}), 400
        
    # Prevent removing privileges from Admin role
    if int(role_id) == 1:
        return jsonify({'error': 'Cannot modify Administrator role privileges.'}), 400
        
    database.update_role_privileges(int(role_id), can_view, can_add, can_edit, can_delete)
    return jsonify({'success': True, 'message': 'Privileges updated successfully.'})

@app.route('/api/admin/categories/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_category():
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    
    if not name:
        return jsonify({'error': 'Category name is required.'}), 400
        
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
        
    cat_id = database.add_category(name, display_order)
    if cat_id:
        return jsonify({'success': True, 'id': cat_id, 'message': 'Category created successfully.'})
    else:
        return jsonify({'error': 'Category name already exists.'}), 409

@app.route('/api/admin/categories/edit/<int:cat_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_category(cat_id):
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    
    if not name:
        return jsonify({'error': 'Category name is required.'}), 400
        
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
        
    success = database.update_category(cat_id, name, display_order)
    if success:
        return jsonify({'success': True, 'message': 'Category updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update category. Ensure the name is unique.'}), 400

@app.route('/api/admin/categories/delete/<int:cat_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_category(cat_id):
    success = database.delete_category(cat_id)
    if success:
        return jsonify({'success': True, 'message': 'Category deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete category.'}), 400

@app.route('/api/bank_modes', methods=['GET'])
def get_bank_modes_api():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    modes = database.get_bank_modes()
    return jsonify(modes)

@app.route('/api/payment_types', methods=['GET'])
def get_payment_types_api():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    types = database.get_payment_types()
    return jsonify(types)

@app.route('/api/payment_categories', methods=['GET'])
def get_payment_categories_api():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    cats = database.get_payment_categories()
    return jsonify(cats)

# ADMIN ROLE EDIT
@app.route('/api/admin/roles/edit', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_role_name():
    data = request.get_json()
    role_id = data.get('role_id')
    name = data.get('name', '').strip()
    if not role_id or not name:
        return jsonify({'error': 'Role ID and name are required.'}), 400
    success = database.update_role_name(int(role_id), name)
    if success:
        return jsonify({'success': True, 'message': 'Role name updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update role name. Ensure it is unique.'}), 400

# ADMIN BANK MODES CRUD
@app.route('/api/admin/bank_modes/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_bank_mode():
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    bm_id = database.add_bank_mode(name, display_order)
    if bm_id:
        return jsonify({'success': True, 'id': bm_id, 'message': 'Bank Mode created successfully.'})
    else:
        return jsonify({'error': 'Name already exists.'}), 409

@app.route('/api/admin/bank_modes/edit/<int:bm_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_bank_mode(bm_id):
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    success = database.update_bank_mode(bm_id, name, display_order)
    if success:
        return jsonify({'success': True, 'message': 'Bank Mode updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update. Ensure name is unique.'}), 400

@app.route('/api/admin/bank_modes/delete/<int:bm_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_bank_mode(bm_id):
    success = database.delete_bank_mode(bm_id)
    if success:
        return jsonify({'success': True, 'message': 'Bank Mode deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete.'}), 400

# ADMIN PAYMENT TYPES CRUD
@app.route('/api/admin/payment_types/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_payment_type():
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    pt_id = database.add_payment_type(name, display_order)
    if pt_id:
        return jsonify({'success': True, 'id': pt_id, 'message': 'Payment Type created successfully.'})
    else:
        return jsonify({'error': 'Name already exists.'}), 409

@app.route('/api/admin/payment_types/edit/<int:pt_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_payment_type(pt_id):
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    success = database.update_payment_type(pt_id, name, display_order)
    if success:
        return jsonify({'success': True, 'message': 'Payment Type updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update. Ensure name is unique.'}), 400

@app.route('/api/admin/payment_types/delete/<int:pt_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_payment_type(pt_id):
    success = database.delete_payment_type(pt_id)
    if success:
        return jsonify({'success': True, 'message': 'Payment Type deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete.'}), 400

# ADMIN PAYMENT CATEGORIES CRUD
@app.route('/api/admin/payment_categories/create', methods=['POST'])
@require_privilege('can_admin')
def admin_create_payment_category():
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    pc_id = database.add_payment_category(name, display_order)
    if pc_id:
        return jsonify({'success': True, 'id': pc_id, 'message': 'Payment Category created successfully.'})
    else:
        return jsonify({'error': 'Name already exists.'}), 409

@app.route('/api/admin/payment_categories/edit/<int:pc_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_payment_category(pc_id):
    data = request.get_json()
    name = data.get('name', '').strip()
    display_order = data.get('display_order', 0)
    if not name:
        return jsonify({'error': 'Name is required.'}), 400
    try:
        display_order = int(display_order)
    except ValueError:
        display_order = 0
    success = database.update_payment_category(pc_id, name, display_order)
    if success:
        return jsonify({'success': True, 'message': 'Payment Category updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update. Ensure name is unique.'}), 400

@app.route('/api/admin/payment_categories/delete/<int:pc_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_payment_category(pc_id):
    success = database.delete_payment_category(pc_id)
    if success:
        return jsonify({'success': True, 'message': 'Payment Category deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete.'}), 400

# YEAR-BASED DEBIT & CREDIT TOTALS API
@app.route('/api/year_totals', methods=['GET'])
def get_year_totals():
    if not is_logged_in():
        return jsonify({'error': 'Unauthorized'}), 401
    year = request.args.get('year')
    if not year:
        return jsonify({'debit': 0.0, 'credit': 0.0})
    
    conn = database.get_db_connection()
    cursor = conn.cursor()
    debit = cursor.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id = ? AND strftime('%Y', date) = ? AND (payment_method = 'Debit' OR payment_method IS NULL OR payment_method = '')",
        (session['user_id'], year)
    ).fetchone()[0] or 0.0
    
    credit = cursor.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id = ? AND strftime('%Y', date) = ? AND payment_method = 'Credit'",
        (session['user_id'], year)
    ).fetchone()[0] or 0.0
    conn.close()
    
    return jsonify({'year': year, 'debit': debit, 'credit': credit})

# CURRENCY CONFIGURATION API
@app.route('/api/active-currency', methods=['GET'])
def api_get_active_currency():
    curr = database.get_active_currency()
    return jsonify(curr)

@app.route('/api/admin/currencies', methods=['GET'])
@require_privilege('can_admin')
def admin_get_currencies():
    currs = database.get_all_currencies()
    return jsonify(currs)

@app.route('/api/admin/currencies/add', methods=['POST'])
@require_privilege('can_admin')
def admin_add_currency():
    data = request.get_json() or {}
    country = data.get('country', '').strip()
    country_desc = data.get('country_desc', '').strip()
    symbol = data.get('symbol', '').strip()
    
    if not country or not country_desc or not symbol:
        return jsonify({'error': 'Country, description, and symbol are required.'}), 400
        
    res_id = database.add_currency(country, country_desc, symbol)
    if res_id is not None:
        return jsonify({'success': True, 'id': res_id, 'message': 'Currency added successfully.'})
    else:
        return jsonify({'error': 'Failed to add currency. Ensure country name is unique.'}), 400

@app.route('/api/admin/currencies/edit/<int:curr_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_edit_currency(curr_id):
    data = request.get_json() or {}
    country = data.get('country', '').strip()
    country_desc = data.get('country_desc', '').strip()
    symbol = data.get('symbol', '').strip()
    
    if not country or not country_desc or not symbol:
        return jsonify({'error': 'Country, description, and symbol are required.'}), 400
        
    success = database.update_currency(curr_id, country, country_desc, symbol)
    if success:
        return jsonify({'success': True, 'message': 'Currency updated successfully.'})
    else:
        return jsonify({'error': 'Failed to update. Ensure country name is unique.'}), 400

@app.route('/api/admin/currencies/set_active/<int:curr_id>', methods=['POST'])
@require_privilege('can_admin')
def admin_set_active_currency(curr_id):
    success = database.set_active_currency(curr_id)
    if success:
        return jsonify({'success': True, 'message': 'Active currency changed successfully.'})
    else:
        return jsonify({'error': 'Failed to set active currency.'}), 400

@app.route('/api/admin/currencies/delete/<int:curr_id>', methods=['POST', 'DELETE'])
@require_privilege('can_admin')
def admin_delete_currency(curr_id):
    success = database.delete_currency(curr_id)
    if success:
        return jsonify({'success': True, 'message': 'Currency deleted successfully.'})
    else:
        return jsonify({'error': 'Failed to delete currency.'}), 400

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
